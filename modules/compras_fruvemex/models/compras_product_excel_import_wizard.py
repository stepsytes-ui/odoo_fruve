import base64
from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook

from odoo import _, fields, models
from odoo.exceptions import ValidationError


class ComprasProductExcelImportWizard(models.TransientModel):
    _name = 'compras.product.excel.import.wizard'
    _description = 'Importar Consolidado desde Excel'

    file_data = fields.Binary(string='Archivo Excel', required=True)
    file_name = fields.Char(string='Nombre de archivo')

    _EXPECTED_HEADERS = [
        ('code', {'codigo'}),
        ('name', {'descripcion'}),
        ('location', {'locacion'}),
        ('model_name', {'modelo'}),
        ('serial', {'serial'}),
        ('manufacturer', {'fabricante'}),
        ('classification', {'clasificacion'}),
        ('category_id', {'categoria'}),
        ('subcategory_id', {'sub categoria', 'subcategoria'}),
        ('subcategory_2', {'sub categoria 2', 'subcategoria 2'}),
        ('repair_line_id', {'equipo linea a', 'equipo - linea a reparar', 'equipo - linea a'}),
        ('expiration_date', {'fecha de cad', 'fecha de caducidad'}),
        ('qty_process', {'qty proceso', 'qty - proceso', 'qty-proceso'}),
        ('total_equipment', {'total equipos', 'total'}),
        ('total_qty_process', {'total qty proceso', 'total qty procesos', 'total qty'}),
        ('frequency_use_days', {'frecuencia uso', 'frecuencia uso dias'}),
        ('unit_id', {'udm', 'unidad'}),
        ('unit_price', {'costo unitario', 'precio unitario'}),
        ('total_cost', {'costo total'}),
        ('currency_id', {'moneda'}),
        ('qty_on_hand', {'inventario'}),
        ('coverage_days', {'cobertura'}),
        ('min_qty', {'min'}),
        ('max_qty', {'max'}),
        ('reorder_point', {'punto reorden', 'punto reorder'}),
        ('lead_time_days', {'tiempo entrega', 'tiempo de entrega', 'tiempo de entrega dias'}),
        ('purchase_type', {'tipo compra', 'tipo de compra'}),
        ('vendor_id', {'proveedor principal', 'proveeder principal'}),
        ('monthly_budget', {'presupuesto mensual'}),
    ]

    def action_import_excel(self):
        self.ensure_one()
        if not self.file_data:
            raise ValidationError(_('Debes seleccionar un archivo Excel.'))

        rows = self._read_excel_rows()
        if not rows:
            raise ValidationError(_('El archivo no contiene datos para importar.'))

        header_row = rows[0]
        self._validate_header_order(header_row)

        product_model = self.env['compras.product']
        created_count = 0
        updated_count = 0
        row_errors = []

        for line_number, row in enumerate(rows[1:], start=2):
            if self._is_empty_row(row):
                continue
            try:
                code = self._as_text(self._cell_value(row, 0))
                if not code:
                    raise ValidationError(_('El Código es obligatorio.'))

                vals = self._build_product_vals_from_row(row)
                product = product_model.search([('code', '=', code)], limit=1)

                if product:
                    product.write(vals)
                    updated_count += 1
                else:
                    vals.setdefault('code', code)
                    if not vals.get('name'):
                        raise ValidationError(_('La Descripción es obligatoria para crear un nuevo registro.'))
                    if not vals.get('unit_id'):
                        default_unit = self.env.ref('uom.product_uom_unit', raise_if_not_found=False)
                        if not default_unit:
                            raise ValidationError(_('No se encontró la unidad por defecto para crear el registro.'))
                        vals['unit_id'] = default_unit.id
                    product_model.create(vals)
                    created_count += 1
            except ValidationError as validation_error:
                row_errors.append(_('Fila %(line)s: %(message)s') % {
                    'line': line_number,
                    'message': validation_error.args[0],
                })

        if row_errors:
            raise ValidationError('\n'.join(row_errors))

        summary = _('%(created)s creados, %(updated)s actualizados.') % {
            'created': created_count,
            'updated': updated_count,
        }
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Importación finalizada'),
                'message': summary,
                'type': 'success',
                'sticky': False,
                'next': {'type': 'ir.actions.act_window_close'},
            },
        }

    def _read_excel_rows(self):
        try:
            workbook = load_workbook(filename=BytesIO(base64.b64decode(self.file_data)), data_only=True)
        except Exception as exc:
            raise ValidationError(_('No se pudo leer el archivo Excel: %s') % str(exc)) from exc

        sheet = workbook.active
        return [list(row) for row in sheet.iter_rows(values_only=True)]

    def _validate_header_order(self, header_row):
        if len(header_row) < len(self._EXPECTED_HEADERS):
            raise ValidationError(_('El archivo debe contener al menos %(count)s columnas en el orden definido.') % {
                'count': len(self._EXPECTED_HEADERS),
            })

        normalized_headers = [self._normalize_header_name(value) for value in header_row]

        for index, (field_name, allowed_names) in enumerate(self._EXPECTED_HEADERS):
            normalized_allowed = {self._normalize_header_name(name) for name in allowed_names}
            if normalized_headers[index] not in normalized_allowed:
                expected_values = ', '.join(sorted(allowed_names))
                raise ValidationError(_(
                    'La columna %(position)s no coincide. Se esperaba: %(expected)s. Valor recibido: %(received)s'
                ) % {
                    'position': index + 1,
                    'expected': expected_values,
                    'received': header_row[index] or '',
                })

    def _build_product_vals_from_row(self, row):
        vals = {}

        name = self._as_text(self._cell_value(row, 1))
        if name:
            vals['name'] = name

        location = self._as_text(self._cell_value(row, 2))
        if location:
            vals['location'] = location

        model_name = self._as_text(self._cell_value(row, 3))
        if model_name:
            vals['model_name'] = model_name

        serial = self._as_text(self._cell_value(row, 4))
        if serial:
            vals['serial'] = serial

        manufacturer = self._as_text(self._cell_value(row, 5))
        if manufacturer:
            vals['manufacturer'] = manufacturer

        classification = self._as_text(self._cell_value(row, 6))
        if classification:
            vals['classification'] = classification

        category_name = self._as_text(self._cell_value(row, 7))
        category = self._find_or_create_category(category_name)
        if category:
            vals['category_id'] = category.id

        subcategory_name = self._as_text(self._cell_value(row, 8))
        subcategory = self._find_or_create_subcategory(subcategory_name, category)
        if subcategory:
            vals['subcategory_id'] = subcategory.id

        subcategory_2 = self._as_text(self._cell_value(row, 9))
        if subcategory_2:
            vals['subcategory_2'] = subcategory_2

        repair_line_name = self._as_text(self._cell_value(row, 10))
        repair_line = self._find_or_create_repair_line(repair_line_name)
        if repair_line:
            vals['repair_line_id'] = repair_line.id

        expiration_date = self._to_date(self._cell_value(row, 11))
        if expiration_date:
            vals['expiration_date'] = expiration_date

        qty_process = self._to_float(self._cell_value(row, 12), _('Qty - Proceso'))
        if qty_process is not None:
            vals['qty_process'] = qty_process

        total_equipment = self._to_float(self._cell_value(row, 13), _('Total Equipos'))
        if total_equipment is not None:
            vals['total_equipment'] = total_equipment

        frequency_use_days = self._to_float(self._cell_value(row, 15), _('Frecuencia Uso'))
        if frequency_use_days is not None:
            vals['frequency_use_days'] = frequency_use_days

        unit_name = self._as_text(self._cell_value(row, 16))
        unit = self._find_uom(unit_name)
        if unit:
            vals['unit_id'] = unit.id

        unit_price = self._to_float(self._cell_value(row, 17), _('Costo Unitario'))
        if unit_price is not None:
            vals['unit_price'] = unit_price

        currency_name = self._as_text(self._cell_value(row, 19))
        currency = self._find_currency(currency_name)
        if currency:
            vals['currency_id'] = currency.id

        qty_on_hand = self._to_float(self._cell_value(row, 20), _('Inventario'))
        if qty_on_hand is not None:
            vals['qty_on_hand'] = qty_on_hand

        min_qty = self._to_float(self._cell_value(row, 22), _('MIN'))
        if min_qty is not None:
            vals['min_qty'] = min_qty

        lead_time_days = self._to_float(self._cell_value(row, 25), _('Tiempo Entrega'))
        if lead_time_days is not None:
            vals['lead_time_days'] = lead_time_days

        purchase_type = self._to_purchase_type(self._cell_value(row, 26))
        if purchase_type:
            vals['purchase_type'] = purchase_type

        vendor_name = self._as_text(self._cell_value(row, 27))
        vendor = self._find_vendor(vendor_name)
        if vendor:
            vals['vendor_id'] = vendor.id

        return vals

    def _find_or_create_category(self, name):
        if not name:
            return False
        category_model = self.env['compras.product.category']
        category = category_model.search([('name', '=', name)], limit=1)
        return category or category_model.create({'name': name})

    def _find_or_create_subcategory(self, name, category):
        if not name or not category:
            return False
        subcategory_model = self.env['compras.product.subcategory']
        subcategory = subcategory_model.search([
            ('name', '=', name),
            ('category_id', '=', category.id),
        ], limit=1)
        return subcategory or subcategory_model.create({'name': name, 'category_id': category.id})

    def _find_or_create_repair_line(self, name):
        if not name:
            return False
        repair_line_model = self.env['compras.repair.line']
        repair_line = repair_line_model.search([('name', '=', name)], limit=1)
        return repair_line or repair_line_model.create({'name': name})

    def _find_uom(self, unit_name):
        if not unit_name:
            return False
        uom_model = self.env['uom.uom']
        unit = uom_model.search([('name', '=', unit_name)], limit=1)
        if not unit:
            unit = uom_model.search([('name', 'ilike', unit_name)], limit=1)
        if not unit:
            raise ValidationError(_('No se encontró la unidad de medida: %s') % unit_name)
        return unit

    def _find_currency(self, currency_name):
        if not currency_name:
            return False
        currency_model = self.env['res.currency']
        currency = currency_model.search([('name', '=', currency_name)], limit=1)
        if not currency:
            currency = currency_model.search([('name', 'ilike', currency_name)], limit=1)
        if not currency:
            currency = currency_model.search([('symbol', '=', currency_name)], limit=1)
        if not currency:
            raise ValidationError(_('No se encontró la moneda: %s') % currency_name)
        return currency

    def _find_vendor(self, vendor_name):
        if not vendor_name:
            return False
        partner_model = self.env['res.partner']
        vendor = partner_model.search([('name', '=', vendor_name)], limit=1)
        if not vendor:
            vendor = partner_model.search([('name', 'ilike', vendor_name)], limit=1)
        if not vendor:
            raise ValidationError(_('No se encontró el proveedor principal: %s') % vendor_name)
        return vendor

    def _to_purchase_type(self, value):
        purchase_type = self._normalize_header_name(value)
        if not purchase_type:
            return False
        if purchase_type in ('local',):
            return 'local'
        if purchase_type in ('internacional', 'international'):
            return 'internacional'
        raise ValidationError(_('Tipo de compra inválido: %s') % value)

    def _to_float(self, value, field_label):
        if value in (None, ''):
            return None
        if isinstance(value, (int, float)):
            return float(value)

        text_value = str(value).strip().replace(',', '')
        if not text_value:
            return None

        try:
            return float(text_value)
        except ValueError as exc:
            raise ValidationError(_('%(field)s no es un número válido: %(value)s') % {
                'field': field_label,
                'value': value,
            }) from exc

    def _to_date(self, value):
        if value in (None, ''):
            return False
        if isinstance(value, datetime):
            return fields.Date.to_string(value.date())
        if isinstance(value, date):
            return fields.Date.to_string(value)

        text_value = str(value).strip()
        if not text_value:
            return False

        parsed_date = fields.Date.to_date(text_value)
        if not parsed_date:
            raise ValidationError(_('Formato de fecha inválido: %s') % value)
        return fields.Date.to_string(parsed_date)

    def _normalize_header_name(self, value):
        normalized = self._as_text(value).lower()
        replacements = {
            'á': 'a',
            'é': 'e',
            'í': 'i',
            'ó': 'o',
            'ú': 'u',
            '.': ' ',
            '_': ' ',
            '/': ' ',
            '-': ' ',
        }
        for old_value, new_value in replacements.items():
            normalized = normalized.replace(old_value, new_value)
        return ' '.join(normalized.split())

    def _cell_value(self, row, index):
        if index >= len(row):
            return None
        return row[index]

    def _is_empty_row(self, row):
        return not any(self._as_text(value) for value in row)

    def _as_text(self, value):
        if value is None:
            return ''
        return str(value).strip()
