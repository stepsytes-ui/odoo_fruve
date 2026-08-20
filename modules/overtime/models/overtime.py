from odoo import api, fields, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import datetime, timedelta
from pytz import timezone
from io import BytesIO
import base64
import re

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None
    Font = PatternFill = Border = Side = Alignment = None
    get_column_letter = None

class Overtime(models.Model):
    _name = 'overtime'
    _description = 'Solicitud de Tiempo Extra'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    
    name = fields.Char(string='Folio', required=True, copy=False, readonly=True, default=lambda self: _('Nuevo'))
    company_id = fields.Many2one('res.company', string='Compañía', required=True, default=lambda self: self.env.company)
    create_date = fields.Datetime(string='Fecha Creación', readonly=True )
    employee_id = fields.Many2one('hr.employee', string='Empleado')
    supervisor_id = fields.Many2one('res.users',string='Supervisor que solicita',default=lambda self: self.env.user.id,readonly=True)
    requested_date = fields.Date(string='Fecha Solicitada', required=True)
    department_id = fields.Many2one('hr.department',string='Departamento',store=True)
    area_id = fields.Many2one('hr.area',string='Área',required=True,domain="[('department_id', '=', department_id)]" )
    justification = fields.Text(string='Explicación del Tiempo Extra', required=True)
    employee_line_ids = fields.One2many('overtime.employee.line', 'overtime_id',string=' ')
    time_from = fields.Float(string='Desde')
    time_to = fields.Float(string='Hasta')
    hours_taken = fields.Float(string='Horas Totales')
    activity = fields.Text(string='Actividad')
    biometric_id = fields.Char(string='Numero de empleado',related='employee_id.biometric_id', store=True)
    attendance_log = fields.Text(string='Registro de checadas del día', compute='_compute_attendance_log', store=False, readonly=True)
    authorized_by_id = fields.Many2one('res.users', string='Autorizado por', readonly=True)
    rejection_reason = fields.Text(string='Motivo de Rechazo', readonly=True)

    state = fields.Selection([
        ('draft', 'Borrador'),
        ('pending', 'Pendiente'),
        ('approved', 'Aprobado'),
        ('rejected', 'Rechazado'),
    ], 
        string='Estado', 
        default='draft', 
        tracking=True
    )

    @api.model
    def _get_overtime_amounts_by_record(self, start_date, end_date, overtimes):
        """Calcula monto por registro aplicando doble o triple segun el empleado.

        La regla de hora triple se evalua sobre las hours_taken de cada
        registro individual (no de forma acumulada): las primeras 9 horas
        se pagan dobles y el excedente, si lo hay, se paga triple.
        """
        amounts_by_record = {}
        if not overtimes:
            return amounts_by_record

        employee_ids = overtimes.mapped('employee_id').ids

        triple_employee_ids = set(
            self.env['overtime.triple.employee'].search([
                ('active', '=', True),
                ('employee_id', 'in', employee_ids),
            ]).mapped('employee_id').ids
        )

        for overtime in overtimes:
            hours = overtime.hours_taken or 0.0
            daily_rate = overtime.employee_id.daily_rate or 0.0
            hourly_rate = daily_rate / 8 if daily_rate else 0.0

            if overtime.employee_id.id in triple_employee_ids:
                double_hours = min(hours, 9.0)
                triple_hours = max(0.0, hours - 9.0)
                amount = (hourly_rate * 2 * double_hours) + (hourly_rate * 3 * triple_hours)
            else:
                amount = hourly_rate * 2 * hours
            amounts_by_record[overtime.id] = amount

        return amounts_by_record

    @api.model
    def get_overtime_dashboard_stats(self, start_date, end_date):
        """
        Calcula las estadísticas del dashboard de tiempo extra
        """
        domain = [
            ('requested_date', '>=', start_date),
            ('requested_date', '<=', end_date),
            ('state', '=', 'approved'),
        ]
        
        overtimes = self.search(domain, order='employee_id, requested_date, id')
        amounts_by_record = self._get_overtime_amounts_by_record(start_date, end_date, overtimes)
        
        # Calcular estadísticas
        total_employees = len(overtimes.mapped('employee_id'))
        total_hours = sum(overtimes.mapped('hours_taken')) or 0
        
        total_pay = sum(amounts_by_record.values())
        
        return {
            'total_employees': total_employees,
            'total_hours': total_hours,
            'total_play': total_pay,
            'total_pay': total_pay,
        }

    @api.model
    def get_overtime_table_data(self, start_date, end_date):
        """
        Retorna los datos para la tabla dinámica del dashboard con columnas por día
        Estructura: {
            'headers': [{'date': '2026-01-16', 'day_label': 'V16'}, ...],
            'rows': [
                {
                    'employee_number': '1234',
                    'employee_name': 'Juan Perez',
                    'daily_rate': 250.00,
                    'days': {
                        '2026-01-16': {'hours': 2.5, 'amount': 125.00},
                        '2026-01-17': {'hours': 0, 'amount': 0},
                    },
                    'total_hours': 2.5,
                    'total_amount': 125.00
                }
            ]
        }
        """
        from datetime import datetime, timedelta
        
        domain = [
            ('requested_date', '>=', start_date),
            ('requested_date', '<=', end_date),
            ('state', '=', 'approved'),
        ]
        
        overtimes = self.search(domain, order='employee_id, requested_date, id')
        amounts_by_record = self._get_overtime_amounts_by_record(start_date, end_date, overtimes)
        
        # Generar todas las fechas del rango
        start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        all_dates = []
        current = start_dt
        while current <= end_dt:
            all_dates.append(current)
            current += timedelta(days=1)
        
        # Encontrar qué días tienen al menos una entrada de tiempo extra
        dates_with_overtime = set()
        for overtime in overtimes:
            dates_with_overtime.add(overtime.requested_date)
        
        # Filtrar solo los días que tienen tiempo extra
        active_dates = [d for d in all_dates if d in dates_with_overtime]
        
        # Construir headers con formato de día (ej: "V16" para viernes 16)
        day_names = ['L', 'M', 'X', 'J', 'V', 'S', 'D']  # Lun, Mar, Mie, Jue, Vie, Sab, Dom
        headers = []
        for date in active_dates:
            day_of_week = day_names[date.weekday()]
            day_number = date.day
            headers.append({
                'date': str(date),
                'day_label': f'{day_of_week}{day_number:02d}'
            })
        
        # Agrupar por empleado y fecha
        employees_data = {}
        
        for overtime in overtimes:
            emp_id = overtime.employee_id.id
            emp_name = overtime.employee_id.name
            emp_biometric = overtime.employee_id.biometric_id or 'N/A'
            emp_daily_rate = overtime.employee_id.daily_rate or 0.0
            overtime_date = str(overtime.requested_date)
            
            if emp_id not in employees_data:
                employees_data[emp_id] = {
                    'employee_number': emp_biometric,
                    'employee_name': emp_name,
                    'daily_rate': emp_daily_rate,
                    'days': {},
                    'total_hours': 0.0,
                    'total_amount': 0.0,
                }
            
            overtime_hours = overtime.hours_taken or 0.0
            overtime_amount = amounts_by_record.get(overtime.id, 0.0)

            # Si ya existe entrada para este día, sumar (en caso de múltiples registros)
            if overtime_date in employees_data[emp_id]['days']:
                employees_data[emp_id]['days'][overtime_date]['hours'] += overtime_hours
                employees_data[emp_id]['days'][overtime_date]['amount'] += overtime_amount
            else:
                employees_data[emp_id]['days'][overtime_date] = {
                    'hours': overtime_hours,
                    'amount': overtime_amount,
                }
        
        # Construir las filas finales
        rows = []
        for emp_id, data in employees_data.items():
            # Asegurar que todos los días activos estén presentes (incluso si es 0)
            for date in active_dates:
                date_str = str(date)
                if date_str not in data['days']:
                    data['days'][date_str] = {'hours': 0.0, 'amount': 0.0}
            
            # Calcular totales
            total_hours = sum(day_data['hours'] for day_data in data['days'].values())
            total_amount = sum(day_data['amount'] for day_data in data['days'].values())
            
            data['total_hours'] = total_hours
            data['total_amount'] = total_amount
            rows.append(data)
        
        # Calcular totales por columna (día)
        column_totals = {}
        grand_total_hours = 0.0
        grand_total_amount = 0.0
        
        for date in active_dates:
            date_str = str(date)
            column_totals[date_str] = {'hours': 0.0, 'amount': 0.0}
            
            for row in rows:
                if date_str in row['days']:
                    column_totals[date_str]['hours'] += row['days'][date_str]['hours']
                    column_totals[date_str]['amount'] += row['days'][date_str]['amount']
            
            grand_total_hours += column_totals[date_str]['hours']
            grand_total_amount += column_totals[date_str]['amount']
        
        # Ordenar filas por número de empleado (menor a mayor)
        rows_sorted = sorted(rows, key=lambda x: int(x['employee_number']) if x['employee_number'].isdigit() else 999999)
        
        return {
            'headers': headers,
            'rows': rows_sorted,
            'column_totals': column_totals,
            'grand_total_hours': grand_total_hours,
            'grand_total_amount': grand_total_amount,
        }

    @api.model
    def export_overtime_table_excel(self, start_date, end_date):
        if Workbook is None:
            raise UserError(_('openpyxl no está instalado. Instale la dependencia para exportar Excel.'))

        table_data = self.get_overtime_table_data(start_date, end_date)
        headers = table_data.get('headers', [])
        rows = table_data.get('rows', [])
        column_totals = table_data.get('column_totals', {})
        grand_total_hours = table_data.get('grand_total_hours', 0.0)
        grand_total_amount = table_data.get('grand_total_amount', 0.0)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Tiempo Extra'
        sheet.sheet_view.showGridLines = False
        safe_start_date = re.sub(r'[^0-9A-Za-z_-]', '_', str(start_date or 'inicio'))
        safe_end_date = re.sub(r'[^0-9A-Za-z_-]', '_', str(end_date or 'fin'))

        fixed_headers = ['No. Empleado', 'Nombre', 'Salario Diario']
        day_headers = []
        for header in headers:
            day_label = header.get('day_label', '')
            day_headers.extend([f'{day_label} Hora', f'{day_label} $'])
        final_headers = fixed_headers + day_headers + ['Total Hrs', 'Total $']

        purple_fill = PatternFill(fill_type='solid', fgColor='8E6AA8')
        blue_fill = PatternFill(fill_type='solid', fgColor='D9EAF7')
        grey_fill = PatternFill(fill_type='solid', fgColor='D9D9D9')
        green_fill = PatternFill(fill_type='solid', fgColor='2E9B3E')
        dark_grey_fill = PatternFill(fill_type='solid', fgColor='C7C7C7')
        white_fill = PatternFill(fill_type='solid', fgColor='FFFFFF')

        border = Border(
            left=Side(style='thin', color='C0C0C0'),
            right=Side(style='thin', color='C0C0C0'),
            top=Side(style='thin', color='C0C0C0'),
            bottom=Side(style='thin', color='C0C0C0'),
        )

        def apply_cell_style(cell, fill=None, font=None, align='center', num_format=None, bold=False):
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font
            else:
                cell.font = Font(bold=bold, color='000000', size=11)
            cell.border = border
            cell.alignment = Alignment(horizontal=align, vertical='center')
            if num_format:
                cell.number_format = num_format

        header_row = 1
        for col_index, header_name in enumerate(final_headers, start=1):
            cell = sheet.cell(row=header_row, column=col_index, value=header_name)
            if col_index <= len(fixed_headers):
                fill = grey_fill
            elif col_index > len(final_headers) - 2:
                fill = green_fill
            else:
                fill = blue_fill
            font = Font(color='000000', bold=True, size=11)
            if col_index > len(final_headers) - 2:
                font = Font(color='FFFFFF', bold=True, size=11)
            apply_cell_style(cell, fill=fill, font=font, align='center', bold=True)
            column_letter = get_column_letter(col_index)
            sheet.column_dimensions[column_letter].width = 16 if col_index != 2 else 26

        row_index = 2
        for row in rows:
            values = [
                row.get('employee_number', ''),
                row.get('employee_name', ''),
                row.get('daily_rate', 0.0),
            ]
            for header in headers:
                date_key = header.get('date')
                day_data = (row.get('days') or {}).get(date_key, {})
                values.extend([day_data.get('hours', 0.0), day_data.get('amount', 0.0)])
            values.extend([row.get('total_hours', 0.0), row.get('total_amount', 0.0)])

            for col_index, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_index, column=col_index, value=value)
                if col_index in (1, 2, 3):
                    fill = white_fill
                    align = 'left' if col_index == 2 else 'center'
                    if col_index == 3:
                        cell.number_format = '$#,##0.00'
                elif col_index > len(final_headers) - 2:
                    fill = green_fill
                    align = 'center' if col_index == len(final_headers) - 1 else 'right'
                    cell.number_format = '$#,##0.00' if col_index == len(final_headers) else '0.00" h"'
                else:
                    fill = white_fill
                    align = 'center' if col_index % 2 == 1 else 'right'
                    cell.number_format = '0.00" h"' if col_index % 2 == 1 else '$#,##0.00'
                apply_cell_style(cell, fill=fill, font=Font(color='000000', bold=False), align=align)
            row_index += 1

        total_values = ['TOTALES'] + [''] * max(len(fixed_headers) - 1, 0)
        for header in headers:
            date_key = header.get('date')
            day_total = column_totals.get(date_key, {})
            total_values.extend([day_total.get('hours', 0.0), day_total.get('amount', 0.0)])
        total_values.extend([grand_total_hours, grand_total_amount])

        for col_index, value in enumerate(total_values, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            if col_index == 1:
                fill = dark_grey_fill
                font = Font(color='000000', bold=True, size=11)
                align = 'center'
            elif col_index > len(final_headers) - 2:
                fill = green_fill
                font = Font(color='FFFFFF', bold=True, size=11)
                align = 'center' if col_index == len(final_headers) - 1 else 'right'
                cell.number_format = '$#,##0.00' if col_index == len(final_headers) else '0.00" h"'
            else:
                fill = blue_fill
                font = Font(color='000000', bold=True, size=11)
                align = 'center' if col_index % 2 == 1 else 'right'
                cell.number_format = '0.00" h"' if col_index % 2 == 1 else '$#,##0.00'
            apply_cell_style(cell, fill=fill, font=font, align=align)

        # Ajuste de anchos y formato final
        for col in sheet.columns:
            first_cell = None
            for cell in col:
                if hasattr(cell, 'column_letter'):
                    first_cell = cell
                    break
            if first_cell is None:
                continue

            max_length = 0
            column = first_cell.column_letter
            for cell in col:
                try:
                    if hasattr(cell, 'value') and cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = max(12, min(max_length + 2, 20))
            sheet.column_dimensions[column].width = adjusted_width

        output = BytesIO()
        try:
            workbook.save(output)
            output.seek(0)
            return {
                'file_name': f'tiempo_extra_{safe_start_date}_{safe_end_date}.xlsx',
                'file_content': base64.b64encode(output.getvalue()).decode('utf-8'),
            }
        finally:
            output.close()

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            # Asignar folio solo cuando el estado es pending
            if vals.get('state') == 'pending':
                # Asegurarse de que la secuencia existe
                sequence = self.env['ir.sequence'].search([('code', '=', 'overtime')], limit=1)
                if not sequence:
                    # Crear la secuencia si no existe
                    sequence = self.env['ir.sequence'].create({
                        'name': 'Solicitud de Tiempo Extra',
                        'code': 'overtime',
                        'prefix': 'OT/%(y)s/%(month)s/',
                        'padding': 4,
                    })
                vals['name'] = sequence.next_by_code('overtime') or _('Nuevo')
        return super().create(vals_list)

    def _assign_folio_if_pending(self):
        """Método para asignar folio cuando el registro está en estado pending"""
        for record in self:
            if record.state == 'pending' and (not record.name or record.name == _('Nuevo')):
                # Asegurarse de que la secuencia existe
                sequence = self.env['ir.sequence'].search([('code', '=', 'overtime')], limit=1)
                if not sequence:
                    # Crear la secuencia si no existe
                    sequence = self.env['ir.sequence'].create({
                        'name': 'Solicitud de Tiempo Extra',
                        'code': 'overtime',
                        'prefix': 'OT/%(y)s/%(month)s/',
                        'padding': 4,
                    })
                record.name = sequence.next_by_code('overtime') or _('Nuevo')

    def write(self, vals):
        result = super().write(vals)
        # Asignar folio después del write si el estado cambió a pending
        if vals.get('state') == 'pending':
            self._assign_folio_if_pending()
        return result

    def action_submit_and_split(self):
        self.ensure_one()
        new_requests = self.env['overtime']

        if not self.employee_line_ids:
            raise ValidationError("Debe agregar al menos un empleado para la solicitud de tiempo extra.")

        # Iterar sobre cada línea de empleado para crear un registro principal individual
        for line in self.employee_line_ids:
            new_vals = {
                'supervisor_id': self.supervisor_id.id,
                'department_id': self.department_id.id,
                'area_id': self.area_id.id,
                'requested_date': self.requested_date,
                'justification': self.justification,
                'state': 'pending',
                'employee_id': line.employee_id.id,
                'time_from': line.time_from,
                'time_to': line.time_to,
                'hours_taken': line.hours_taken,
                'activity': line.activity,
            }

            new_request = new_requests.create(new_vals)
            new_requests |= new_request

        # Asegurar que todos los registros tengan folio asignado
        new_requests._assign_folio_if_pending()

        # Eliminar la solicitud original (la que tiene las líneas)
        self.unlink()

        return {
            'name': _('Solicitudes de Tiempo Extra Creadas'),
            'res_model': 'overtime',
            'view_mode': 'list,form',
            'domain': [('id', 'in', new_requests.ids)],
            'target': 'main',
            'type': 'ir.actions.act_window',
        }

    def action_approve(self):
        self.ensure_one()
        if self.state not in ('pending', 'rejected'):
            raise ValidationError(_('Solo se pueden aprobar solicitudes en estado Pendiente o Rechazado.'))
        self.authorized_by_id = self.env.user.id
        self.state = 'approved'

    def action_reject(self):
        self.ensure_one()
        if self.state not in ('pending', 'approved'):
            raise ValidationError(_('Solo se pueden rechazar solicitudes en estado Pendiente o Aprobado.'))
        return {
            'name': 'Motivo de Rechazo',
            'type': 'ir.actions.act_window',
            'res_model': 'overtime.rejection.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {'default_overtime_id': self.id}
        }
    
    @api.depends('employee_id', 'requested_date')
    def _compute_attendance_log(self):
        Attendance = self.env['hr.attendance']
        tz_name = self.env.context.get('tz') or self.env.user.tz or 'UTC'
        user_tz = timezone(tz_name)

        for record in self:
            record.attendance_log = False

            if record.employee_id and record.requested_date:
                date_start = fields.Datetime.to_datetime(record.requested_date)
                date_end = date_start + timedelta(days=2)

                attendances = Attendance.search([
                    ('employee_id', '=', record.employee_id.id),
                    ('check_in', '>=', date_start),
                    ('check_in', '<', date_end),
                ], order='check_in asc')

                log_lines=[]
                if attendances:
                    for att in attendances:
                        if att.check_in:
                            check_in_time = fields.Datetime.context_timestamp(att, timestamp=att.check_in)
                            formatted_time = check_in_time.strftime('%d-%m-%Y %H:%M:%S')
                            log_lines.append(formatted_time)
                    
                    record.attendance_log = '\n'.join(log_lines)
                else:
                    record.attendance_log = "No se han registrado checadas del empleado"
