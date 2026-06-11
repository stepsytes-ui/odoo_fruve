from odoo import models, fields, api, _
from datetime import datetime, timedelta, time
import pytz
import logging
from io import BytesIO
import base64
import html as _html

_logger = logging.getLogger(__name__)

FIXED_DEVICE_TIMEZONE_NAME = 'America/Tijuana'

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    _logger.warning("openpyxl not installed. Excel generation will not work.")
    Workbook = None


class AttendanceReportWizard(models.TransientModel):
    _name = 'attendance.report.wizard'
    _description = 'Wizard para Generar Reporte de Asistencia Personalizado'

    company_id = fields.Many2one(
        'res.company',
        string='Empresa',
        required=True,
        readonly=True,
        default=lambda self: self.env.company,
    )
    date_from = fields.Date(string='Fecha Desde', required=True)
    date_to = fields.Date(string='Fecha Hasta', required=True)
    employee_id = fields.Many2one(
        'hr.employee',
        string='Empleado (Opcional)',
        domain="[('company_id', '=', company_id)]",
        help='Dejar en blanco para incluir todos los empleados'
    )
    show_archived = fields.Boolean(
        string='Ver Empleados Archivados',
        default=False,
        help='Activa esta opción para ver la asistencia de empleados que ya fueron dados de baja (finiquitados)'
    )

    @api.constrains('date_from', 'date_to')
    def _check_dates(self):
        for record in self:
            if record.date_from > record.date_to:
                raise ValueError(_('La fecha "Desde" no puede ser posterior a la fecha "Hasta"'))

    def action_preview_report(self):
        """Abre una vista previa HTML del reporte en una nueva pestaña del navegador"""
        return {
            'type': 'ir.actions.act_url',
            'url': f'/attendance/report/preview/{self.id}',
            'target': 'new',
        }

    def action_generate_report(self):
        """Descarga el reporte en Excel directamente sin crear archivos adjuntos"""
        return {
            'type': 'ir.actions.act_url',
            'url': f'/attendance/report/download/{self.id}',
            'target': 'self',
        }

    def _get_weekly_period_info(self, date_from, date_to):
        """Retorna info de semana custom (viernes-jueves) si el rango coincide exactamente."""
        if not date_from or not date_to:
            return None

        if date_to != (date_from + timedelta(days=6)):
            return None

        # Semana custom inicia viernes (4) y termina jueves (3)
        if date_from.weekday() != 4 or date_to.weekday() != 3:
            return None

        iso_year, iso_week, _ = date_to.isocalendar()
        return {
            'custom_year': iso_year,
            'custom_week': iso_week,
            'week_start_date': date_from,
            'week_end_date': date_to,
        }

    def _build_late_weekly_form_url(self, report_id):
        base_url = self.env.company.get_attendance_reports_base_url()
        action = self.env.ref('zkteco_realtime_connector.action_attendance_late_weekly_report', raise_if_not_found=False)
        action_part = f"&action={action.id}" if action else ""
        return f"{base_url}/web#id={report_id}&model=attendance.late.weekly.report&view_type=form{action_part}"

    @api.model
    def _cron_send_weekly_attendance_email(self):
        """Envia correo semanal (viernes) con Excel de asistencia por empresa."""
        group = self.env.ref(
            'zkteco_realtime_connector.group_receive_weekly_attendance_email',
            raise_if_not_found=False,
        )
        if not group:
            _logger.warning('[WEEKLY ATTENDANCE EMAIL] Grupo de destinatarios no encontrado.')
            return {'sent': 0, 'companies': 0, 'skipped': 'group_not_found'}

        companies = self.env['res.company'].sudo().search([])
        sent_mails = 0
        processed_companies = 0

        for company in companies:
            tz_name = company.timezone or FIXED_DEVICE_TIMEZONE_NAME
            try:
                company_tz = pytz.timezone(tz_name)
            except pytz.UnknownTimeZoneError:
                company_tz = pytz.timezone(FIXED_DEVICE_TIMEZONE_NAME)

            now_local = datetime.now(company_tz)
            # Tomar siempre la ultima semana completa viernes-jueves.
            # Si hoy es jueves, usa el jueves de la semana anterior para evitar periodos incompletos.
            days_since_thursday = (now_local.weekday() - 3) % 7
            if days_since_thursday == 0:
                days_since_thursday = 7
            week_end = now_local.date() - timedelta(days=days_since_thursday)
            week_start = week_end - timedelta(days=6)
            iso_year, iso_week, iso_weekday = week_end.isocalendar()

            recipients = self.env['res.users'].sudo().search([
                ('active', '=', True),
                ('share', '=', False),
                ('groups_id', 'in', group.id),
                ('company_ids', 'in', company.id),
                ('partner_id.email', '!=', False),
            ])

            if not recipients:
                _logger.info(
                    '[WEEKLY ATTENDANCE EMAIL] Sin destinatarios para empresa %s.',
                    company.display_name,
                )
                continue

            wizard = self.sudo().create({
                'company_id': company.id,
                'date_from': week_start,
                'date_to': week_end,
                'show_archived': False,
            })

            try:
                excel_bytes, filename = wizard._generate_excel_file()
            except Exception as error:
                _logger.error(
                    '[WEEKLY ATTENDANCE EMAIL] Error generando Excel para %s: %s',
                    company.display_name,
                    error,
                    exc_info=True,
                )
                continue

            attachment = self.env['ir.attachment'].sudo().create({
                'name': filename,
                'type': 'binary',
                'datas': base64.b64encode(excel_bytes),
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            })

            date_from_txt = week_start.strftime('%d/%m/%Y')
            date_to_txt = week_end.strftime('%d/%m/%Y')
            subject = _('ASISTENCIA SEMANA %(week)s - %(start)s al %(end)s') % {
                'week': iso_week,
                'start': date_from_txt,
                'end': date_to_txt,
            }
            body = _(
                '<p>Hola,</p>'
                '<p>Adjuntamos el reporte de asistencia semanal de Odoo.</p>'
                '<p>Periodo: <strong>%(start)s</strong> al <strong>%(end)s</strong> (Semana %(week)s).</p>'
                '<p>Saludos.</p>'
            ) % {
                'start': date_from_txt,
                'end': date_to_txt,
                'week': iso_week,
            }

            mail_values = {
                'subject': subject,
                'body_html': body,
                'recipient_ids': [(4, user.partner_id.id) for user in recipients if user.partner_id],
                'attachment_ids': [(4, attachment.id)],
                'email_from': company.email or self.env.user.email_formatted or 'odoo@localhost',
                'auto_delete': False,
            }
            self.env['mail.mail'].sudo().create(mail_values).send()

            sent_mails += 1
            processed_companies += 1
            _logger.info(
                '[WEEKLY ATTENDANCE EMAIL] Correo semanal enviado para %s (semana %s).',
                company.display_name,
                iso_week,
            )

        return {'sent': sent_mails, 'companies': processed_companies}

    def _get_report_data(self):
        """
        Construye y retorna la estructura de datos del reporte.
        Usada tanto por la vista previa HTML como por la generación de Excel.
        """
        date_from = self.date_from
        date_to = self.date_to
        report_company = self.company_id or self.env.company

        company_tz_name = report_company.timezone or FIXED_DEVICE_TIMEZONE_NAME
        try:
            COMPANY_TZ = pytz.timezone(company_tz_name)
        except pytz.UnknownTimeZoneError:
            COMPANY_TZ = pytz.timezone(FIXED_DEVICE_TIMEZONE_NAME)

        if self.employee_id and self.employee_id.company_id != report_company:
            raise ValueError(
                _('El empleado seleccionado no pertenece a la empresa activa: %s')
                % report_company.display_name
            )

        if self.show_archived:
            employee_domain = [
                ('company_id', '=', report_company.id),
                ('active', '=', False),
                ('finiquitado', '=', True),
                ('turno_id', '!=', False),
            ]
        else:
            employee_domain = [
                ('company_id', '=', report_company.id),
                '|',
                ('employee_status', '=', 'active'),
                '&',
                ('employee_status', '=', 'inactive'),
                ('finiquitado', '=', False),
                ('turno_id', '!=', False)
            ]
        if self.employee_id:
            employee_domain.append(('id', '=', self.employee_id.id))

        if self.show_archived:
            employees = self.env['hr.employee'].with_context(active_test=False).search(employee_domain)
        else:
            employees = self.env['hr.employee'].search(employee_domain)

        try:
            employees = sorted(employees, key=lambda e: int(e.biometric_id) if e.biometric_id else 0)
        except (ValueError, TypeError):
            employees = sorted(employees, key=lambda e: e.biometric_id or '')

        if not employees:
            raise ValueError(_('No se encontraron empleados con los criterios especificados.'))

        employee_ids = [emp.id for emp in employees]

        date_list = []
        current_date = date_from
        while current_date <= date_to:
            date_list.append(current_date)
            current_date += timedelta(days=1)

        day_names = {
            0: 'Lunes', 1: 'Martes', 2: 'Miércoles', 3: 'Jueves',
            4: 'Viernes', 5: 'Sábado', 6: 'Domingo',
        }

        weekly_info = self._get_weekly_period_info(date_from, date_to)
        show_neto_total = bool(weekly_info)

        late_weekly_map = {}
        observations_map = {}
        if show_neto_total:
            weekly_reports = self.env['attendance.late.weekly.report'].search([
                ('company_id', '=', report_company.id),
                ('custom_year', '=', weekly_info['custom_year']),
                ('custom_week', '=', weekly_info['custom_week']),
                ('employee_id', 'in', employee_ids),
            ])
            late_weekly_map = {rec.employee_id.id: rec for rec in weekly_reports if rec.employee_id}

            weekly_adjustments = self.env['attendance.late.weekly.adjustment'].search([
                ('employee_id', 'in', employee_ids),
                ('custom_year', '=', weekly_info['custom_year']),
                ('custom_week', '=', weekly_info['custom_week']),
            ])
            observations_map = {
                rec.employee_id.id: (rec.manual_observation or '')
                for rec in weekly_adjustments
                if rec.employee_id
            }

        Attendance = self.env['hr.attendance']
        rows = []
        for employee in employees:
            cells = {}
            for date_obj in date_list:
                cells[date_obj] = self._get_cell_data_for_employee_date(
                    employee, date_obj, COMPANY_TZ, Attendance
                )

            payable_days = None
            payable_days_url = ''
            observation = ''
            if show_neto_total:
                weekly_record = late_weekly_map.get(employee.id)
                if weekly_record:
                    payable_days = float(weekly_record.payable_days or 0.0)
                    payable_days_url = self._build_late_weekly_form_url(weekly_record.id)
                else:
                    # Sin retardos/faltas en la semana: neto completo de 6 dias.
                    payable_days = 6.0
                observation = observations_map.get(employee.id, '')

            rows.append({
                'biometric_id': employee.biometric_id or '',
                'name': employee.name or '',
                'turno': employee.sudo().turno_id.turno_name or '',
                'cells': cells,
                'payable_days': payable_days,
                'payable_days_url': payable_days_url,
                'employee_id': employee.id,
                'custom_year': weekly_info['custom_year'] if show_neto_total else None,
                'custom_week': weekly_info['custom_week'] if show_neto_total else None,
                'observation': observation,
            })

        return {
            'date_list': date_list,
            'day_names': day_names,
            'rows': rows,
            'date_from': date_from,
            'date_to': date_to,
            'company': report_company,
            'company_tz': COMPANY_TZ,
            'show_neto_total': show_neto_total,
        }

    def _build_preview_html(self):
        """Genera el HTML completo de la vista previa del reporte"""
        data = self._get_report_data()
        date_list = data['date_list']
        day_names = data['day_names']
        rows = data['rows']
        date_from = data['date_from']
        date_to = data['date_to']
        show_neto_total = data.get('show_neto_total', False)

        def cell_style(color, font_color, bold):
            styles = [
                'padding:4px 6px',
                'border:1px solid #ccc',
                'white-space:pre-line',
                'font-size:11px',
                'vertical-align:top',
                'line-height:1.4',
            ]
            if color:
                styles.append(f'background-color:#{color}')
            if font_color:
                styles.append(f'color:#{font_color}')
            if bold:
                styles.append('font-weight:bold')
            return '; '.join(styles)

        def esc(val):
            return _html.escape(str(val or ''))

        parts = []
        parts.append(f'''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Vista Previa Asistencia {esc(str(date_from))} – {esc(str(date_to))}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: Arial, sans-serif; font-size: 13px; background: #f0f2f5; }}
  .toolbar {{
    position: sticky; top: 0; z-index: 200;
    background: #fff; padding: 8px 16px;
    border-bottom: 2px solid #4472C4;
    display: flex; align-items: center; gap: 12px;
    box-shadow: 0 2px 6px rgba(0,0,0,0.12);
  }}
  .toolbar h2 {{ flex: 1; font-size: 14px; color: #1a2b4a; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }}
  .toolbar .meta {{ font-size: 11px; color: #6c757d; white-space: nowrap; }}
  .btn {{ padding: 6px 16px; border: none; border-radius: 4px; cursor: pointer; font-size: 13px; text-decoration: none; display: inline-block; font-weight: 600; }}
  .btn-primary {{ background: #4472C4; color: #fff; }}
  .btn-primary:hover {{ background: #2e57a8; }}
  .legend {{
    background: #fff; padding: 6px 16px; position: sticky; top: 44px; z-index: 190;
    border-bottom: 1px solid #dde;
    display: flex; flex-wrap: wrap; gap: 14px; align-items: center;
    font-size: 11px; color: #444;
  }}
  .legend-item {{ display: flex; align-items: center; gap: 5px; }}
  .lswatch {{ width: 13px; height: 13px; border-radius: 2px; border: 1px solid rgba(0,0,0,0.2); display: inline-block; flex-shrink: 0; }}
  .table-wrap {{ overflow: auto; max-height: calc(100vh - 92px); }}
  table {{ border-collapse: collapse; background: #fff; min-width: 100%; }}
  th, td {{ border: 1px solid #ccc; font-size: 11px; }}
  th {{
    background: #4472C4; color: #fff;
    padding: 5px 7px; white-space: nowrap;
    position: sticky; top: 0; z-index: 10;
    text-align: center;
  }}
  td {{ padding: 4px 6px; vertical-align: top; }}
  .s0 {{ position: sticky; left: 0;   z-index: 5; background: inherit; min-width: 75px;  max-width: 75px; }}
  .s1 {{ position: sticky; left: 75px; z-index: 5; background: inherit; min-width: 170px; max-width: 170px; }}
  .s2 {{ position: sticky; left: 245px; z-index: 5; background: inherit; min-width: 110px; max-width: 110px; }}
    .s3 {{ position: sticky; left: 355px; z-index: 5; background: inherit; min-width: 95px; max-width: 95px; text-align: center; box-shadow: 2px 0 0 0 #ccc; }}
    .obs-col {{ min-width: 170px; max-width: 260px; }}
    th.s0, th.s1, th.s2, th.s3 {{ z-index: 20; }}
    th.s0, th.s1, th.s2, th.s3 {{ background-color: #4472C4 !important; color: #FFFFFF !important; }}
    th.s3 {{ box-shadow: 2px 0 0 0 #2a53a0 !important; }}
  .day-cell {{ min-width: 95px; max-width: 130px; }}
    .neto-link {{ color: #1b5fbd; font-weight: 700; text-decoration: none; }}
    .neto-link:hover {{ text-decoration: underline; }}
    .obs-link {{ color: #1b5fbd; text-decoration: none; font-weight: 600; }}
    .obs-link:hover {{ text-decoration: underline; }}
    td.is-active-row {{
        box-shadow: inset 0 1px 0 rgba(96, 96, 96, 0.75), inset 0 -1px 0 rgba(96, 96, 96, 0.75);
    }}
    td.is-active-col {{
        box-shadow: inset 1px 0 0 rgba(96, 96, 96, 0.75), inset -1px 0 0 rgba(96, 96, 96, 0.75);
    }}
    td.is-active-row.is-active-col {{
        box-shadow:
            inset 0 1px 0 rgba(96, 96, 96, 0.75),
            inset 0 -1px 0 rgba(96, 96, 96, 0.75),
            inset 1px 0 0 rgba(96, 96, 96, 0.75),
            inset -1px 0 0 rgba(96, 96, 96, 0.75);
    }}
    td.is-active-cell {{
        outline: 2px solid #1b5fbd;
        outline-offset: -2px;
        background-color: #dce8ff !important;
        color: #000 !important;
    }}
  tr:nth-child(even) td {{ background-color: #f9f9fb; }}
    tr:hover td {{ background-color: #eef3ff !important; color: #000 !important; }}
    tr:hover td * {{ color: #000 !important; }}
    td:hover {{ background-color: #dce8ff !important; color: #000 !important; }}
    td:hover * {{ color: #000 !important; }}
    .col-hidden {{ display: none !important; }}
    .col-menu {{
        position: fixed;
        z-index: 500;
        min-width: 230px;
        background: #fff;
        border: 1px solid #cfd5e3;
        box-shadow: 0 8px 24px rgba(0, 0, 0, 0.18);
        border-radius: 6px;
        padding: 6px;
        display: none;
    }}
    .col-menu button {{
        width: 100%;
        border: none;
        background: transparent;
        text-align: left;
        padding: 8px 10px;
        font-size: 12px;
        border-radius: 4px;
        cursor: pointer;
    }}
    .col-menu button:hover {{ background: #eef3ff; }}
    .col-menu .sep {{ border-top: 1px solid #e2e7f0; margin: 6px 0; }}
</style>
</head>
<body>
<div class="toolbar">
  <h2>&#128203; Vista Previa &mdash; Reporte de Asistencia</h2>
  <span class="meta">{esc(str(date_from))} al {esc(str(date_to))} &bull; {len(rows)} empleado(s)</span>
  <a class="btn btn-primary" href="/attendance/report/download/{self.id}">&#11015; Descargar Excel</a>
</div>
<div class="legend">
  <strong>Leyenda:</strong>
  <div class="legend-item"><span class="lswatch" style="background:#00B050"></span>Asistencia</div>
  <div class="legend-item"><span class="lswatch" style="background:#FF0000"></span>Falta</div>
  <div class="legend-item"><span class="lswatch" style="background:#FFC000"></span>Incapacidad</div>
  <div class="legend-item"><span class="lswatch" style="background:#FFC7CE"></span><span style="color:#FF6600">Permiso</span></div>
  <div class="legend-item"><span class="lswatch" style="background:#4472C4"></span>Festivo</div>
  <div class="legend-item"><span class="lswatch" style="background:#FF6600"></span>En finiquito</div>
  <div class="legend-item"><span class="lswatch" style="background:#fff"></span><span style="color:#808080">Descanso</span></div>
</div>
<div class="table-wrap">
<table id="attendance-preview-table">
<thead><tr>
    <th class="s0" data-hideable="1" data-col-label="No. Emp.">No. Emp.</th>
    <th class="s1" data-hideable="1" data-col-label="Nombre">Nombre</th>
    <th class="s2" data-hideable="1" data-col-label="Turno">Turno</th>
''')

        if show_neto_total:
                        parts.append('  <th class="s3" data-hideable="1" data-col-label="Neto Total">Neto Total</th>\n')

        for date_obj in date_list:
            day_name = day_names[date_obj.weekday()]
            parts.append(
                f'  <th class="day-cell" data-hideable="1" data-col-label="{esc(day_name)} {date_obj.strftime("%d/%m")}">{esc(day_name)}<br>'
                f'<span style="font-weight:normal;font-size:10px">{date_obj.strftime("%d/%m")}</span></th>\n'
            )

        if show_neto_total:
            parts.append('  <th class="obs-col" data-hideable="1" data-col-label="Observaciones">Observaciones</th>\n')

        parts.append('</tr></thead>\n<tbody>\n')

        for row in rows:
            parts.append('<tr>\n')
            parts.append(f'  <td class="s0">{esc(row["biometric_id"])}</td>\n')
            parts.append(f'  <td class="s1">{esc(row["name"])}</td>\n')
            parts.append(f'  <td class="s2">{esc(row["turno"])}</td>\n')
            if show_neto_total:
                value = row.get('payable_days')
                if value is None:
                    parts.append('  <td class="s3">-</td>\n')
                else:
                    neto_text = f"{value:.2f}"
                    neto_url = row.get('payable_days_url')
                    if neto_url:
                        parts.append(
                            f'  <td class="s3"><a class="neto-link" href="{esc(neto_url)}" target="_blank">{esc(neto_text)}</a></td>\n'
                        )
                    else:
                        parts.append(f'  <td class="s3"><strong>{esc(neto_text)}</strong></td>\n')
            for date_obj in date_list:
                cell = row['cells'].get(date_obj, {})
                style = cell_style(cell.get('color'), cell.get('font_color'), cell.get('bold', False))
                text = esc(cell.get('text', '') or '')
                parts.append(f'  <td class="day-cell" style="{style}">{text}</td>\n')

            if show_neto_total:
                obs_text = (row.get('observation') or '').strip()
                obs_label = obs_text[:40] + ('...' if len(obs_text) > 40 else '') if obs_text else 'Agregar'
                parts.append(
                    f'  <td class="obs-col">'
                    f'<a href="#" class="obs-link" '
                    f'data-employee-id="{row.get("employee_id")}" '
                    f'data-custom-year="{row.get("custom_year")}" '
                    f'data-custom-week="{row.get("custom_week")}" '
                    f'data-observation="{esc(obs_text)}">{esc(obs_label)}</a>'
                    f'</td>\n'
                )
            parts.append('</tr>\n')

        parts.append(f'''</tbody>
</table>
</div>
<script>
(() => {{
    const table = document.getElementById('attendance-preview-table');
    const hiddenColumns = new Map();
    let activeCell = null;

    const menu = document.createElement('div');
    menu.className = 'col-menu';
    menu.id = 'column-context-menu';
    document.body.appendChild(menu);
    let contextColumnIndex = -1;

    function isCellVisible(cell) {{
        return !!cell && !cell.classList.contains('col-hidden');
    }}

    function getBodyRows() {{
        return table ? Array.from(table.querySelectorAll('tbody tr')) : [];
    }}

    function getColumnLabel(index) {{
        if (!table) return '';
        const th = table.querySelectorAll('thead th')[index];
        if (!th) return `Columna ${{index + 1}}`;
        return th.dataset.colLabel || th.textContent.trim().replace(/\s+/g, ' ') || `Columna ${{index + 1}}`;
    }}

    function hideMenu() {{
        menu.style.display = 'none';
    }}

    function renderMenu() {{
        const actions = [];
        const currentLabel = getColumnLabel(contextColumnIndex);
        actions.push(
            `<button type="button" data-action="hide-current">Ocultar columna: ${{currentLabel}}</button>`
        );

        if (hiddenColumns.size) {{
            actions.push('<div class="sep"></div>');
            actions.push('<button type="button" data-action="show-all">Mostrar todas</button>');
            hiddenColumns.forEach((label, index) => {{
                actions.push(
                    `<button type="button" data-action="show-one" data-index="${{index}}">Mostrar: ${{label}}</button>`
                );
            }});
        }}

        menu.innerHTML = actions.join('');
    }}

    function setColumnHidden(index, hidden) {{
        if (!table || index < 0) return;
        const rows = table.querySelectorAll('tr');
        rows.forEach((row) => {{
            const cell = row.children[index];
            if (cell) cell.classList.toggle('col-hidden', hidden);
        }});

        if (hidden) {{
            hiddenColumns.set(index, getColumnLabel(index));
            if (activeCell && activeCell.cellIndex === index) {{
                clearCrosshair();
                activeCell = null;
            }}
        }} else {{
            hiddenColumns.delete(index);
        }}

        updateStickyOffsets();
    }}

    function updateStickyOffsets() {{
        if (!table) return;

        const stickyClasses = ['s0', 's1', 's2', 's3'];
        const stickyHeaders = Array.from(
            table.querySelectorAll('thead th.s0, thead th.s1, thead th.s2, thead th.s3')
        );
        let left = 0;
        let lastVisibleIndex = -1;

        stickyHeaders.forEach((headerCell) => {{
            const isHidden = headerCell.classList.contains('col-hidden');
            if (!isHidden) lastVisibleIndex = headerCell.cellIndex;
        }});

        stickyHeaders.forEach((headerCell) => {{
            const colIndex = headerCell.cellIndex;
            const isHidden = headerCell.classList.contains('col-hidden');
            const isLastSticky = colIndex === lastVisibleIndex;

            table.querySelectorAll('tr').forEach((row) => {{
                const cell = row.children[colIndex];
                if (!cell) return;
                if (!stickyClasses.some((cls) => cell.classList.contains(cls))) return;
                cell.style.left = isHidden ? '' : `${{left}}px`;
                const isHeader = cell.tagName === 'TH';
                if (isLastSticky && !isHidden) {{
                    cell.style.boxShadow = isHeader ? '2px 0 0 0 #2a53a0' : '2px 0 0 0 #ccc';
                }} else {{
                    cell.style.boxShadow = '';
                }}
            }});

            if (!isHidden) {{
                left += headerCell.offsetWidth;
            }}
        }});
    }}

    function focusCell(cell) {{
        if (!cell || !isCellVisible(cell)) return;
        applyCrosshair(cell);
        activeCell = cell;
        cell.focus({{ preventScroll: true }});
    }}

    function findNearestVisibleCell(row, preferredIndex) {{
        if (!row) return null;
        if (isCellVisible(row.children[preferredIndex])) return row.children[preferredIndex];
        for (let distance = 1; distance < row.children.length; distance += 1) {{
            const left = preferredIndex - distance;
            const right = preferredIndex + distance;
            if (left >= 0 && isCellVisible(row.children[left])) return row.children[left];
            if (right < row.children.length && isCellVisible(row.children[right])) return row.children[right];
        }}
        return null;
    }}

    function findHorizontalCell(row, fromIndex, direction) {{
        if (!row) return null;
        let index = fromIndex + direction;
        while (index >= 0 && index < row.children.length) {{
            if (isCellVisible(row.children[index])) return row.children[index];
            index += direction;
        }}
        return null;
    }}

    function clearCrosshair() {{
        if (!table) return;
        table.querySelectorAll('td.is-active-row, td.is-active-col, td.is-active-cell').forEach((td) => {{
            td.classList.remove('is-active-row', 'is-active-col', 'is-active-cell');
        }});
    }}

    function applyCrosshair(cell) {{
        if (!table || !cell) return;

        const row = cell.parentElement;
        if (!row) return;

        const rowCells = Array.from(row.children);
        const colIndex = rowCells.indexOf(cell);
        if (colIndex < 0) return;

        clearCrosshair();

        rowCells.forEach((td) => td.classList.add('is-active-row'));
        table.querySelectorAll('tbody tr').forEach((tr) => {{
            const td = tr.children[colIndex];
            if (isCellVisible(td)) td.classList.add('is-active-col');
        }});
        cell.classList.add('is-active-cell');
    }}

    if (table) {{
        updateStickyOffsets();

        table.querySelectorAll('tbody td').forEach((td) => {{
            td.tabIndex = -1;
        }});

        table.addEventListener('click', (ev) => {{
            const cell = ev.target.closest('td');
            if (!cell || !table.contains(cell)) return;
            if (!isCellVisible(cell)) return;
            focusCell(cell);
        }});

        table.addEventListener('contextmenu', (ev) => {{
            const th = ev.target.closest('thead th[data-hideable="1"]');
            if (!th || !table.contains(th)) return;
            ev.preventDefault();

            contextColumnIndex = th.cellIndex;
            renderMenu();
            menu.style.left = `${{ev.clientX}}px`;
            menu.style.top = `${{ev.clientY}}px`;
            menu.style.display = 'block';
        }});

        table.addEventListener('keydown', (ev) => {{
            if (!activeCell) return;

            const key = ev.key;
            if (!['ArrowLeft', 'ArrowRight', 'ArrowUp', 'ArrowDown'].includes(key)) return;
            ev.preventDefault();

            const rows = getBodyRows();
            const row = activeCell.parentElement;
            const rowIndex = rows.indexOf(row);
            const colIndex = activeCell.cellIndex;
            let target = null;

            if (key === 'ArrowLeft') {{
                target = findHorizontalCell(row, colIndex, -1);
            }}
            if (key === 'ArrowRight') {{
                target = findHorizontalCell(row, colIndex, 1);
            }}
            if (key === 'ArrowUp' || key === 'ArrowDown') {{
                const step = key === 'ArrowUp' ? -1 : 1;
                let nextIndex = rowIndex + step;
                while (nextIndex >= 0 && nextIndex < rows.length) {{
                    target = findNearestVisibleCell(rows[nextIndex], colIndex);
                    if (target) break;
                    nextIndex += step;
                }}
            }}

            if (target) focusCell(target);
        }});
    }}

    menu.addEventListener('click', (ev) => {{
        const button = ev.target.closest('button[data-action]');
        if (!button) return;

        const action = button.dataset.action;
        if (action === 'hide-current') {{
            setColumnHidden(contextColumnIndex, true);
        }}
        if (action === 'show-all') {{
            Array.from(hiddenColumns.keys()).forEach((index) => setColumnHidden(index, false));
        }}
        if (action === 'show-one') {{
            const index = parseInt(button.dataset.index || '-1', 10);
            setColumnHidden(index, false);
        }}
        hideMenu();
    }});

    document.addEventListener('click', (ev) => {{
        if (!menu.contains(ev.target)) hideMenu();
    }});

    document.addEventListener('scroll', hideMenu, true);
    window.addEventListener('resize', hideMenu);

    async function saveObservation(payload) {{
        const response = await fetch('/attendance/report/save_weekly_observation', {{
            method: 'POST',
            headers: {{ 'Content-Type': 'application/json' }},
            body: JSON.stringify(payload),
        }});
        if (!response.ok) {{
            throw new Error('No se pudo guardar observacion');
        }}
        return response.json();
    }}

    document.addEventListener('click', async (ev) => {{
        const link = ev.target.closest('.obs-link');
        if (!link) return;
        ev.preventDefault();

        const current = link.dataset.observation || '';
        const value = window.prompt('Observaciones de la semana:', current);
        if (value === null) return;

        const payload = {{
            wizard_id: {self.id},
            employee_id: parseInt(link.dataset.employeeId || '0', 10),
            custom_year: parseInt(link.dataset.customYear || '0', 10),
            custom_week: parseInt(link.dataset.customWeek || '0', 10),
            observation: value,
        }};

        try {{
            const result = await saveObservation(payload);
            if (result && result.ok) {{
                const saved = result.observation || '';
                const label = saved ? (saved.length > 40 ? saved.slice(0, 40) + '...' : saved) : 'Agregar';
                link.textContent = label;
                link.dataset.observation = saved;
            }}
        }} catch (err) {{
            window.alert('Error guardando observacion.');
        }}
    }});
}})();
</script>
</body>
</html>''')
        return ''.join(parts)

    def _generate_excel_file(self):
        """Genera el archivo Excel y retorna (bytes_data, filename) sin crear adjuntos."""
        if Workbook is None:
            raise ImportError('openpyxl no está instalado. Instálelo con: pip install openpyxl')

        data = self._get_report_data()
        date_list = data['date_list']
        day_names = data['day_names']
        rows = data['rows']
        date_from = data['date_from']
        date_to = data['date_to']
        show_neto_total = data.get('show_neto_total', False)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Reporte de Asistencia'

        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        fixed_headers = ['No. Empleado', 'Nombre', 'Turno']
        if show_neto_total:
            fixed_headers.append('Neto Total')
        col_num = 1
        for header in fixed_headers:
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
            col_num += 1

        date_columns = {}
        for date_obj in date_list:
            day_name = day_names[date_obj.weekday()]
            header_text = f'{day_name} {date_obj.day}'
            cell = ws.cell(row=1, column=col_num, value=header_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
            date_columns[date_obj] = col_num
            col_num += 1

        observation_col = None
        if show_neto_total:
            observation_col = col_num
            cell = ws.cell(row=1, column=observation_col, value='Observaciones')
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
            col_num += 1

        total_cols = col_num

        row_num = 2
        for row in rows:
            ws.cell(row=row_num, column=1, value=row['biometric_id'])
            ws.cell(row=row_num, column=2, value=row['name'])
            ws.cell(row=row_num, column=3, value=row['turno'])

            start_fixed_cols = 3
            if show_neto_total:
                start_fixed_cols = 4
                neto_cell = ws.cell(row=row_num, column=4, value=row['payable_days'])
                neto_cell.number_format = '0.00'
                if row.get('payable_days_url'):
                    neto_cell.hyperlink = row['payable_days_url']
                    neto_cell.style = 'Hyperlink'

            for col in range(1, start_fixed_cols + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            for date_obj, col_idx in date_columns.items():
                cell_data = row['cells'].get(date_obj, {})
                cell = ws.cell(row=row_num, column=col_idx, value=cell_data.get('text', '') or '')
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                if cell_data.get('color'):
                    cell.fill = PatternFill(
                        start_color=cell_data['color'],
                        end_color=cell_data['color'],
                        fill_type='solid'
                    )
                if cell_data.get('font_color'):
                    cell.font = Font(color=cell_data['font_color'], bold=cell_data.get('bold', False))

            if show_neto_total and observation_col:
                obs_cell = ws.cell(row=row_num, column=observation_col, value=(row.get('observation') or ''))
                obs_cell.border = border
                obs_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            row_num += 1

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        date_col_start = 4
        if show_neto_total:
            ws.column_dimensions['D'].width = 12
            date_col_start = 5

        for col in range(date_col_start, total_cols):
            ws.column_dimensions[get_column_letter(col)].width = 25

        if show_neto_total and observation_col:
            ws.column_dimensions[get_column_letter(observation_col)].width = 35

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'Reporte_Asistencia_{date_from}_{date_to}.xlsx'
        return output.getvalue(), filename

    def _format_hour_from_float(self, hour_value):
        """Convierte hora float (ej. 14.5) a HH:MM."""
        if hour_value is False or hour_value is None:
            return None
        hours = int(hour_value)
        minutes = int(round((hour_value - hours) * 60))
        if minutes == 60:
            hours += 1
            minutes = 0
        hours = hours % 24
        return f"{hours:02d}:{minutes:02d}"

    def _get_attendance_punches_text(self, employee, start_utc_str, end_utc_str, company_tz, Attendance):
        """Retorna las checadas del día (check_in) como texto separado por guiones."""
        punches = Attendance.search([
            ('employee_id', '=', employee.id),
            ('check_in', '>=', start_utc_str),
            ('check_in', '<=', end_utc_str),
            ('punctuality_status', '!=', 'absence')
        ], order='check_in asc')

        time_list = []
        for att in punches:
            check_in_dt = att.check_in
            if check_in_dt.tzinfo is None:
                check_in_dt = pytz.utc.localize(check_in_dt)
            local_dt = check_in_dt.astimezone(company_tz)
            time_list.append(local_dt.strftime('%H:%M:%S'))

        return ' - '.join(time_list)

    def _build_hourly_leave_cell_text(self, leave_record, employee, start_utc_str, end_utc_str, company_tz, Attendance):
        """Construye texto para permiso por horas y agrega checadas del mismo día."""
        hour_from = self._format_hour_from_float(leave_record.request_hour_from)
        hour_to = self._format_hour_from_float(leave_record.request_hour_to)

        if not hour_from or not hour_to:
            date_from = leave_record.date_from
            date_to = leave_record.date_to
            if date_from:
                if date_from.tzinfo is None:
                    date_from = pytz.utc.localize(date_from)
                hour_from = date_from.astimezone(company_tz).strftime('%H:%M')
            if date_to:
                if date_to.tzinfo is None:
                    date_to = pytz.utc.localize(date_to)
                hour_to = date_to.astimezone(company_tz).strftime('%H:%M')

        leave_title = f"Permiso por horas {hour_from or '--:--'} - {hour_to or '--:--'}"
        punches_text = self._get_attendance_punches_text(
            employee, start_utc_str, end_utc_str, company_tz, Attendance
        )

        if punches_text:
            return f"{leave_title}\n{punches_text}"
        return leave_title

    def _get_cell_data_for_employee_date(self, employee, target_date, company_tz, Attendance):
        """
        Retorna información completa para una celda: texto, color de fondo y color de fuente.
        Considera: check-ins, descanso, falta, vacaciones y todos los tipos de permiso (NEW_LEAVE_STATUS).
        Para Incapacidad, ignora los descansos del shift_management.
        """
        try:
            COMPANY_TZ = company_tz
        except:
            COMPANY_TZ = pytz.timezone(FIXED_DEVICE_TIMEZONE_NAME)
        
        # Crear rango UTC para la fecha
        start_of_day_local = COMPANY_TZ.localize(datetime.combine(target_date, time.min))
        end_of_day_local = COMPANY_TZ.localize(datetime.combine(target_date, time.max))

        start_of_day_utc = start_of_day_local.astimezone(pytz.utc)
        end_of_day_utc = end_of_day_local.astimezone(pytz.utc)

        start_utc_str = fields.Datetime.to_string(start_of_day_utc)
        end_utc_str = fields.Datetime.to_string(end_of_day_utc)

        descanso_statuses = ['leave_day_off', 'Descanso', 'descanso']
        incapacidad_statuses = ['leave_sickness', 'leave_sickness_paid', 'Incapacidad', 'incapacidad']


        CalendarLeaves = self.env['resource.calendar.leaves']
        midday_local = COMPANY_TZ.localize(datetime.combine(target_date, time(12, 0, 0)))
        midday_utc_str = fields.Datetime.to_string(midday_local.astimezone(pytz.utc))
        public_holiday = CalendarLeaves.search([
            ('resource_id', '=', False),  # Festivo global (aplica a todos)
            ('date_from', '<=', midday_utc_str),
            ('date_to', '>=', midday_utc_str),
        ], limit=1)

        if public_holiday:
            return {
                'text': public_holiday.name,
                'color': '4472C4',  # Azul
                'font_color': 'FFFFFF',  # Blanco
                'bold': True
            }
        
        if not employee.sudo().turno_id:
            return {'text': '', 'color': None, 'font_color': None, 'bold': False}
        
        shift = employee.sudo().turno_id
        
        # Manejo especial para turno "ESPECIAL" (gerentes/dueños)
        if shift.turno_name == 'ESPECIAL':
            # Respetar Incapacidad/Descanso registrados en attendance.
            sickness_attendance = Attendance.search([
                ('employee_id', '=', employee.id),
                ('check_in', '>=', start_utc_str),
                ('check_in', '<=', end_utc_str),
                ('punctuality_status', 'in', incapacidad_statuses)
            ], limit=1)

            if sickness_attendance:
                return {
                    'text': 'Incapacidad',
                    'color': 'FFC000',
                    'font_color': 'FFFFFF',
                    'bold': True
                }

            descanso_attendance = Attendance.search([
                ('employee_id', '=', employee.id),
                ('check_in', '>=', start_utc_str),
                ('check_in', '<=', end_utc_str),
                ('punctuality_status', 'in', descanso_statuses)
            ], limit=1)

            if descanso_attendance:
                return {
                    'text': 'Descanso',
                    'color': None,
                    'font_color': '808080',
                    'bold': False
                }

            # Verificar si hay checadas válidas
            valid_attendances = Attendance.search([
                ('employee_id', '=', employee.id),
                ('check_in', '>=', start_utc_str),
                ('check_in', '<=', end_utc_str),
                ('punctuality_status', 'in', ['on_time', 'late', 'LunchS', 'LunchE', 'end', 'overtime', 'n/a'])
            ], order='check_in asc')
            
            if valid_attendances:
                # Si hay checadas, mostrarlas
                check_in_times = []
                for att in valid_attendances:
                    utc_datetime = pytz.utc.localize(att.check_in)
                    local_datetime = utc_datetime.astimezone(COMPANY_TZ)
                    time_str = local_datetime.strftime("%H:%M:%S")
                    
                    if att.punctuality_status == 'n/a' and att.check_out:
                        utc_checkout = pytz.utc.localize(att.check_out)
                        local_checkout = utc_checkout.astimezone(COMPANY_TZ)
                        checkout_str = local_checkout.strftime("%H:%M:%S")
                        time_str = f"{time_str} - {checkout_str}"
                    
                    check_in_times.append(time_str)
                
                return {
                    'text': ' - '.join(check_in_times),
                    'color': '00B050',  # Verde
                    'font_color': 'FFFFFF',  # Blanco
                    'bold': False
                }
            else:
                # Si no hay checadas, mostrar palomita
                return {
                    'text': '✓',
                    'color': '00B050',  # Verde
                    'font_color': 'FFFFFF',  # Blanco
                    'bold': True
                }
        
        if shift.turno_name == 'Seguridad':

            Leave = self.env['hr.leave'].sudo()
            sickness_leave = Leave.search([
                ('employee_id', '=', employee.id),
                ('state', '=', 'validate'),
                ('holiday_status_id.name', '=', 'Incapacidad'),
                ('date_from', '<=', end_utc_str),
                ('date_to', '>=', start_utc_str)
            ], limit=1)

            sickness_attendance = Attendance.search([
                ('employee_id', '=', employee.id),
                ('check_in', '>=', start_utc_str),
                ('check_in', '<=', end_utc_str),
                ('punctuality_status', 'in', incapacidad_statuses)
            ], limit=1)
            
            if sickness_leave or sickness_attendance:
                return {
                    'text': 'Incapacidad',
                    'color': 'FFC000',  
                    'font_color': 'FFFFFF',  
                    'bold': True
                }

            # Para Seguridad, respetar faltas capturadas manualmente por RH.
            # El cron de faltas no genera ausencias para Seguridad/ESPECIAL.
            absence_record = Attendance.search([
                ('employee_id', '=', employee.id),
                ('check_in', '>=', start_utc_str),
                ('check_in', '<=', end_utc_str),
                ('punctuality_status', '=', 'absence')
            ], limit=1)

            if absence_record:
                return {
                    'text': 'Falta',
                    'color': 'FF0000',
                    'font_color': 'FFFFFF',
                    'bold': True
                }

            descanso_attendance = Attendance.search([
                ('employee_id', '=', employee.id),
                ('check_in', '>=', start_utc_str),
                ('check_in', '<=', end_utc_str),
                ('punctuality_status', 'in', descanso_statuses)
            ], limit=1)

            if descanso_attendance:
                return {
                    'text': 'Descanso',
                    'color': None,
                    'font_color': '808080',
                    'bold': False
                }
            
            approved_leave = Leave.search([
                ('employee_id', '=', employee.id),
                ('state', '=', 'validate'),
                ('date_from', '<=', end_utc_str),
                ('date_to', '>=', start_utc_str)
            ], limit=1)
            
            if approved_leave:
                leave_name = approved_leave.holiday_status_id.name
                if approved_leave.request_unit_hours:
                    hourly_text = self._build_hourly_leave_cell_text(
                        approved_leave, employee, start_utc_str, end_utc_str, COMPANY_TZ, Attendance
                    )
                    return {
                        'text': hourly_text,
                        'color': 'FFC7CE',
                        'font_color': 'FF6600',
                        'bold': True
                    }
                if leave_name.strip().lower() == 'descanso':
                    return {
                        'text': 'Descanso',
                        'color': None,
                        'font_color': '808080',  # Gris
                        'bold': False
                    }
                return {
                    'text': leave_name,
                    'color': 'FFC7CE',  # Rosa claro
                    'font_color': 'FF6600',  # Naranja oscuro
                    'bold': True
                }
            
            # 5a. Si hay check-ins válidos, mostrarlos
            valid_attendances = Attendance.search([
                ('employee_id', '=', employee.id),
                ('check_in', '>=', start_utc_str),
                ('check_in', '<=', end_utc_str),
                ('punctuality_status', 'in', ['on_time', 'late', 'LunchS', 'LunchE', 'end', 'overtime','n/a'])
            ], order='check_in asc')
            
            if valid_attendances:
                check_in_times = []
                for att in valid_attendances:
                    utc_datetime = pytz.utc.localize(att.check_in)
                    local_datetime = utc_datetime.astimezone(COMPANY_TZ)
                    time_str = local_datetime.strftime("%H:%M:%S")
                    
                    # Si es 'n/a', incluir también check_out si existe
                    if att.punctuality_status == 'n/a' and att.check_out:
                        utc_checkout = pytz.utc.localize(att.check_out)
                        local_checkout = utc_checkout.astimezone(COMPANY_TZ)
                        checkout_str = local_checkout.strftime("%H:%M:%S")
                        time_str = f"{time_str} / {checkout_str}"
                    
                    check_in_times.append(time_str)
                
                return {
                    'text': ' - '.join(check_in_times),
                    'color': '00B050',  # Verde
                    'font_color': 'FFFFFF',  # Blanco
                    'bold': False
                }
            
            # Si no hay check-ins, mostrar Descanso
            return {
                'text': 'Descanso',
                'color': None,
                'font_color': '808080',  # Gris
                'bold': False
            }
        
        day_mapping = {
            0: 'work_monday',
            1: 'work_tuesday',
            2: 'work_wednesday',
            3: 'work_thursday',
            4: 'work_friday',
            5: 'work_saturday',
            6: 'work_sunday',
        }
        day_of_week = target_date.weekday()
        field_to_check = day_mapping.get(day_of_week)
        
        # Verificar si el día es laboral (considerando también días especiales)
        is_work_day = getattr(shift, field_to_check, False)
        
        # 2. Verificar Incapacidad PRIMERO (ignora descansos del shift)
        Leave = self.env['hr.leave'].sudo()
        sickness_leave = Leave.search([
            ('employee_id', '=', employee.id),
            ('state', '=', 'validate'),
            ('holiday_status_id.name', '=', 'Incapacidad'),
            ('date_from', '<=', end_utc_str),
            ('date_to', '>=', start_utc_str)
        ], limit=1)

        sickness_attendance = Attendance.search([
            ('employee_id', '=', employee.id),
            ('check_in', '>=', start_utc_str),
            ('check_in', '<=', end_utc_str),
            ('punctuality_status', 'in', incapacidad_statuses)
        ], limit=1)
        
        if sickness_leave or sickness_attendance:
            return {
                'text': 'Incapacidad',
                'color': 'FFC000',  # Amarillo/Oro
                'font_color': 'FFFFFF',  # Blanco
                'bold': True
            }

        # 3. Verificar Descanso en hr.attendance (si existe, marcar descanso)
        descanso_attendance = Attendance.search([
            ('employee_id', '=', employee.id),
            ('check_in', '>=', start_utc_str),
            ('check_in', '<=', end_utc_str),
            ('punctuality_status', 'in', descanso_statuses)
        ], limit=1)

        if descanso_attendance:
            return {
                'text': 'Descanso',
                'color': None,
                'font_color': '808080',  # Gris
                'bold': False
            }
        
        # 4. Verificar otros tipos de permiso/licencia
        approved_leave = Leave.search([
            ('employee_id', '=', employee.id),
            ('state', '=', 'validate'),
            ('date_from', '<=', end_utc_str),
            ('date_to', '>=', start_utc_str)
        ], limit=1)
        
        if approved_leave:
            leave_name = approved_leave.holiday_status_id.name
            if approved_leave.request_unit_hours:
                hourly_text = self._build_hourly_leave_cell_text(
                    approved_leave, employee, start_utc_str, end_utc_str, COMPANY_TZ, Attendance
                )
                return {
                    'text': hourly_text,
                    'color': 'FFC7CE',
                    'font_color': 'FF6600',
                    'bold': True
                }
            # Descanso desde hr.leave → cuadro blanco, texto gris
            if leave_name.strip().lower() == 'descanso' or day_of_week == 6:
                return {
                    'text': 'Descanso',
                    'color': None,
                    'font_color': '808080',  # Gris
                    'bold': False
                }
            return {
                'text': leave_name,
                'color': 'FFC7CE',  # Rosa claro
                'font_color': 'FF6600',  # Naranja oscuro
                'bold': True
            }
        
        # 4. Buscar attendances
        attendances = Attendance.search([
            ('employee_id', '=', employee.id),
            ('check_in', '>=', start_utc_str),
            ('check_in', '<=', end_utc_str)
        ], order='check_in asc')
        
        # 5. Verificar si hay falta
        absence_record = Attendance.search([
            ('employee_id', '=', employee.id),
            ('check_in', '>=', start_utc_str),
            ('check_in', '<=', end_utc_str),
            ('punctuality_status', '=', 'absence')
        ], limit=1)
        
        if absence_record:
            return {
                'text': 'Falta',
                'color': 'FF0000',  # Rojo
                'font_color': 'FFFFFF',  # Blanco
                'bold': True
            }
        
        # 6. Si hay check-ins válidos, mostrarlos
        valid_attendances = Attendance.search([
            ('employee_id', '=', employee.id),
            ('check_in', '>=', start_utc_str),
            ('check_in', '<=', end_utc_str),
            ('punctuality_status', 'in', ['on_time', 'late', 'LunchS', 'LunchE', 'end', 'overtime', 'n/a'])
        ], order='check_in asc')
        
        if valid_attendances:
            check_in_times = []
            for att in valid_attendances:
                utc_datetime = pytz.utc.localize(att.check_in)
                local_datetime = utc_datetime.astimezone(COMPANY_TZ)
                time_str = local_datetime.strftime("%H:%M:%S")
                
                # Si es 'n/a', incluir también check_out si existe
                if att.punctuality_status == 'n/a' and att.check_out:
                    utc_checkout = pytz.utc.localize(att.check_out)
                    local_checkout = utc_checkout.astimezone(COMPANY_TZ)
                    checkout_str = local_checkout.strftime("%H:%M:%S")
                    time_str = f"{time_str} - {checkout_str}"
                
                check_in_times.append(time_str)
            
            return {
                'text': ' - '.join(check_in_times),
                'color': '00B050',  # Verde
                'font_color': 'FFFFFF',  # Blanco
                'bold': False
            }
        
        # 7. Si NO es un día laboral, mostrar "Descanso"
        if not is_work_day:
            return {
                'text': 'Descanso',
                'color': None,
                'font_color': '808080',  # Gris
                'bold': False
            }
        
        # 8. Si es día laboral sin check-ins y sin falta
        # Aquí sí verificamos si está en proceso de finiquito (solo para días laborales sin asistencia)
        if employee.employee_status == 'inactive' and not employee.finiquitado:
            return {
                'text': 'En proceso de finiquito',
                'color': 'FF6600',  # Naranja fuerte
                'font_color': 'FFFFFF',  # Blanco
                'bold': True
            }
        return {'text': '', 'color': None, 'font_color': None, 'bold': False}

    def _get_check_ins_for_employee_date(self, employee, target_date, company_tz, Attendance):
        """Obtiene todos los check-in times para un empleado en una fecha específica"""
        
        # Crear rango UTC para la fecha
        start_of_day_local = company_tz.localize(datetime.combine(target_date, time.min))
        end_of_day_local = company_tz.localize(datetime.combine(target_date, time.max))

        start_of_day_utc = start_of_day_local.astimezone(pytz.utc)
        end_of_day_utc = end_of_day_local.astimezone(pytz.utc)

        start_utc_str = fields.Datetime.to_string(start_of_day_utc)
        end_utc_str = fields.Datetime.to_string(end_of_day_utc)

        # Buscar attendances
        attendances = Attendance.search([
            ('employee_id', '=', employee.id),
            ('check_in', '>=', start_utc_str),
            ('check_in', '<=', end_utc_str),
            ('punctuality_status', 'in', ['on_time', 'late', 'LunchS', 'LunchE', 'end', 'overtime', 'n/a'])
        ], order='check_in asc')

        check_in_times = []
        for att in attendances:
            # Convertir a hora local
            utc_datetime = pytz.utc.localize(att.check_in)
            local_datetime = utc_datetime.astimezone(company_tz)
            time_str = local_datetime.strftime("%H:%M:%S")
            check_in_times.append(time_str)

        return check_in_times
