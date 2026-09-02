# -*- coding: utf-8 -*-

import base64
import logging
from io import BytesIO

from odoo import _, fields, models
from odoo.exceptions import UserError

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:
    Workbook = None


_logger = logging.getLogger(__name__)


class ExportEmployeeListWizard(models.TransientModel):
    _name = 'export.employee.list.wizard'
    _description = 'Asistente para exportar lista de empleados'

    employee_list_type = fields.Selection(
        [
            ('summary', 'Lista por departamento'),
            ('detailed', 'Lista detallada'),
        ],
        string='Tipo de lista',
        required=True,
        default='detailed',
    )

    department_id = fields.Many2one(
        'hr.department',
        string='Departamento',
    )

    employee_status_filter = fields.Selection(
        [
            ('active', 'Activos'),
            ('inactive', 'Inactivos'),
        ],
        string='Mostrar empleados',
        required=True,
        default='active',
    )

    archivo_excel = fields.Binary(
        string='Archivo Excel',
        readonly=True,
        attachment=False,
    )

    nombre_archivo = fields.Char(
        string='Nombre del Archivo',
        readonly=True,
    )

    def _get_employees(self):
        self.ensure_one()
        domain = [('active', '=', self.employee_status_filter == 'active')]
        if self.employee_list_type == 'summary' and self.department_id:
            domain.append(('department_id', '=', self.department_id.id))
        employees = self.env['hr.employee'].with_context(active_test=False).search(domain)

        def sort_key(employee):
            if self.employee_list_type == 'summary':
                department_name = (employee.department_id.name or '').strip()
                employee_name = (employee.name or '').strip()
                return (
                    not department_name,
                    department_name.casefold(),
                    employee_name.casefold(),
                    employee.id,
                )

            biometric_value = (employee.biometric_id or '').strip()
            if biometric_value.isdigit():
                number_key = (0, int(biometric_value))
            else:
                number_key = (1, biometric_value)
            return number_key + ((employee.name or '').strip(), employee.id)

        return employees.sorted(key=sort_key)

    def _generate_excel_file(self):
        if Workbook is None:
            raise UserError(_('openpyxl no está instalado. Instale la dependencia para exportar Excel.'))

        employees = self._get_employees()

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Lista de Empleados'

        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        if self.employee_list_type == 'summary':
            headers = ['No. Empleado', 'Nombre', 'Departamento']
        else:
            headers = ['No. Empleado', 'Nombre', 'CURP', 'RFC', 'IMSS', 'Turno', 'Estado']
        for column_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=column_index, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border

        for row_index, employee in enumerate(employees, start=2):
            if self.employee_list_type == 'summary':
                row_values = [
                    employee.biometric_id or '',
                    employee.name or '',
                    employee.department_id.name or '',
                ]
            else:
                row_values = [
                    employee.biometric_id or '',
                    employee.name or '',
                    employee.identification_id or '',
                    employee.rfc or '',
                    employee.ssnid or '',
                    employee.turno_id.name or '',
                    'Activo' if employee.active else 'Inactivo',
                ]
            for column_index, value in enumerate(row_values, start=1):
                cell = sheet.cell(row=row_index, column=column_index, value=value)
                cell.alignment = left_alignment
                cell.border = border

        sheet.freeze_panes = 'A2'
        sheet.auto_filter.ref = sheet.dimensions
        column_widths = [16, 32, 24] if self.employee_list_type == 'summary' else [16, 32, 24, 22, 18, 20, 14]
        for column_index, width in enumerate(column_widths, start=1):
            sheet.column_dimensions[chr(64 + column_index)].width = width

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        status_suffix = 'activos' if self.employee_status_filter == 'active' else 'inactivos'
        if self.employee_list_type == 'summary':
            filename = f'lista_datos_empleados_por_departamento_{status_suffix}.xlsx'
        else:
            filename = f'lista_datos_empleados_{status_suffix}.xlsx'
        return output.getvalue(), filename

    def action_export_excel(self):
        self.ensure_one()

        excel_bytes, filename = self._generate_excel_file()
        self.write({
            'archivo_excel': base64.b64encode(excel_bytes),
            'nombre_archivo': filename,
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/?model=export.employee.list.wizard&id={self.id}&field=archivo_excel&filename_field=nombre_archivo&download=true',
            'target': 'self',
        }