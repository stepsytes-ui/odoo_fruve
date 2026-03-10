from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError
from datetime import datetime, timedelta, time
import pytz
import logging
from io import BytesIO
import base64

_logger = logging.getLogger(__name__)

FIXED_DEVICE_TIMEZONE_NAME = 'America/Tijuana'

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    _logger.warning("openpyxl not installed. Excel generation will not work.")
    Workbook = None

try:
    import xlrd
except ImportError:
    _logger.warning("xlrd not installed. XLS file reading will not work.")
    xlrd = None


class AttendanceImportWizard(models.TransientModel):
    _name = 'attendance.import.wizard'
    _description = 'Wizard para Importar Excel de Checadas y Generar Reporte'

    excel_file = fields.Binary(
        string='Archivo Excel',
        required=True,
        help='Archivo Excel con columnas: Número Empleado, Nombre, Departamento, Hora Checada'
    )
    filename = fields.Char(string='Nombre del Archivo')
    date_from = fields.Date(string='Fecha Desde', required=True)
    date_to = fields.Date(string='Fecha Hasta', required=True)

    @api.constrains('date_from', 'date_to')
    def _check_dates(self):
        for record in self:
            if record.date_from > record.date_to:
                raise ValidationError(_('La fecha "Desde" no puede ser posterior a la fecha "Hasta"'))

    def action_generate_report(self):
        """Lee el Excel importado y genera el reporte de asistencias"""
        self.ensure_one()
        
        if not self.excel_file:
            raise UserError(_('Debe cargar un archivo Excel.'))
        
        try:
            COMPANY_TZ = pytz.timezone(FIXED_DEVICE_TIMEZONE_NAME)
        except pytz.UnknownTimeZoneError:
            _logger.error(f"Error: Zona horaria '{FIXED_DEVICE_TIMEZONE_NAME}' es inválida.")
            raise UserError(_('Configuración de zona horaria inválida.'))

        # Decodificar el archivo Excel
        try:
            file_content = base64.b64decode(self.excel_file)
            
            # Detectar si es XLS o XLSX basándose en el nombre del archivo
            is_xls = self.filename and self.filename.lower().endswith('.xls') and not self.filename.lower().endswith('.xlsx')
            
            if is_xls:
                # Leer archivo XLS usando xlrd
                if xlrd is None:
                    raise UserError(_('La librería xlrd no está instalada. Por favor, instale xlrd para leer archivos .xls'))
                
                workbook = xlrd.open_workbook(file_contents=file_content)
                sheet = workbook.sheet_by_index(0)
                
                # Función para iterar sobre las filas de xlrd
                def get_rows():
                    for row_idx in range(1, sheet.nrows):  # Empezar desde 1 para saltar encabezado
                        row = []
                        for col_idx in range(min(4, sheet.ncols)):  # Solo primeras 4 columnas
                            cell = sheet.cell(row_idx, col_idx)
                            
                            # Convertir fechas de Excel a datetime
                            if cell.ctype == 3:  # XL_CELL_DATE
                                try:
                                    date_tuple = xlrd.xldate_as_tuple(cell.value, workbook.datemode)
                                    row.append(datetime(*date_tuple))
                                except:
                                    row.append(cell.value)
                            else:
                                row.append(cell.value)
                        
                        yield row_idx + 1, row
                
                rows_iterator = get_rows()
            else:
                # Leer archivo XLSX usando openpyxl
                workbook = load_workbook(BytesIO(file_content))
                sheet = workbook.active
                rows_iterator = enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2)
                
        except Exception as e:
            raise UserError(_('Error al leer el archivo Excel: %s\n\nSugerencia: Si su archivo es .xls (Excel antiguo), intente guardarlo como .xlsx (Excel moderno) o instale la librería xlrd.') % str(e))

        # Leer datos del Excel
        # Formato: Columna A=Número Empleado, B=Nombre, C=Departamento, D=Hora Checada
        checadas_data = []
        employee_info = {}  # {numero_empleado: {'nombre': '', 'departamento': '', 'turno': ''}}
        
        for row_idx, row in rows_iterator:
            if not row[0]:  # Si no hay número de empleado, saltar
                continue
            
            numero_empleado = str(row[0]).strip()
            nombre = str(row[1]).strip() if row[1] else ''
            departamento = str(row[2]).strip() if row[2] else ''
            hora_checada_str = row[3]
            
            # Guardar info del empleado
            if numero_empleado not in employee_info:
                # Buscar el empleado en Odoo para obtener turno
                employee = self.env['hr.employee'].search([
                    ('biometric_id', '=', numero_empleado)
                ], limit=1)
                
                turno = employee.sudo().turno_id.turno_name if employee and employee.sudo().turno_id else 'N/A'
                
                employee_info[numero_empleado] = {
                    'nombre': nombre,
                    'departamento': departamento,
                    'turno': turno,
                    'employee_record': employee
                }
            
            # Procesar hora de checada
            if hora_checada_str:
                try:
                    # Intentar parsear la hora
                    if isinstance(hora_checada_str, datetime):
                        fecha_hora = hora_checada_str
                    else:
                        # Intentar varios formatos
                        for fmt in ['%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M:%S', '%m/%d/%Y %H:%M:%S', '%Y-%m-%d %H:%M']:
                            try:
                                fecha_hora = datetime.strptime(str(hora_checada_str), fmt)
                                break
                            except ValueError:
                                continue
                        else:
                            _logger.warning(f"No se pudo parsear la fecha/hora en fila {row_idx}: {hora_checada_str}")
                            continue
                    
                    checadas_data.append({
                        'numero_empleado': numero_empleado,
                        'fecha_hora': fecha_hora,
                        'fecha': fecha_hora.date()
                    })
                except Exception as e:
                    _logger.warning(f"Error procesando checada en fila {row_idx}: {e}")
                    continue
        
        if not employee_info:
            raise UserError(_('No se encontraron empleados en el archivo Excel.'))
        
        # Organizar checadas por empleado y fecha
        # {numero_empleado: {fecha: [lista de datetime]}}
        checadas_por_empleado = {}
        for checada in checadas_data:
            emp_num = checada['numero_empleado']
            fecha = checada['fecha']
            
            if emp_num not in checadas_por_empleado:
                checadas_por_empleado[emp_num] = {}
            
            if fecha not in checadas_por_empleado[emp_num]:
                checadas_por_empleado[emp_num][fecha] = []
            
            checadas_por_empleado[emp_num][fecha].append(checada['fecha_hora'])
        
        # Ordenar checadas por hora
        for emp_num in checadas_por_empleado:
            for fecha in checadas_por_empleado[emp_num]:
                checadas_por_empleado[emp_num][fecha].sort()
        
        # Generar lista de fechas para el reporte
        date_list = []
        current_date = self.date_from
        while current_date <= self.date_to:
            date_list.append(current_date)
            current_date += timedelta(days=1)
        
        # Determinar qué días tienen checadas de varios empleados (para detectar días laborales)
        dias_con_checadas = {}  # {fecha: cantidad_de_empleados_con_checada}
        for fecha in date_list:
            count = sum(1 for emp_num in checadas_por_empleado if fecha in checadas_por_empleado[emp_num])
            dias_con_checadas[fecha] = count

        descanso_dates, incapacidad_dates, public_holiday_dates = self._get_special_day_maps(
            employee_info,
            date_list,
            COMPANY_TZ
        )
        
        # Crear workbook de salida
        wb = Workbook()
        ws = wb.active
        ws.title = 'Reporte de Asistencias'

        # Estilos
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Encabezados fijos
        fixed_headers = ['No. Empleado', 'Nombre', 'Departamento', 'Turno']
        col_num = 1
        
        for header in fixed_headers:
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
            col_num += 1

        # Encabezados dinámicos por día
        day_names = {
            0: 'Lunes',
            1: 'Martes',
            2: 'Miércoles',
            3: 'Jueves',
            4: 'Viernes',
            5: 'Sábado',
            6: 'Domingo'
        }
        
        date_columns = {}  # Mapeo de fecha a número de columna
        for date_obj in date_list:
            day_name = day_names[date_obj.weekday()]
            header_text = f'{day_name} {date_obj.day}'
            
            cell = ws.cell(row=1, column=col_num, value=header_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
            
            date_columns[date_obj] = col_num
            col_num += 1

        # Llenar datos de empleados
        row_num = 2
        
        # Ordenar empleados por número
        sorted_employees = sorted(employee_info.items(), key=lambda x: int(x[0]) if x[0].isdigit() else 0)
        
        for numero_empleado, info in sorted_employees:
            ws.cell(row=row_num, column=1, value=numero_empleado)
            ws.cell(row=row_num, column=2, value=info['nombre'])
            ws.cell(row=row_num, column=3, value=info['departamento'])
            ws.cell(row=row_num, column=4, value=info['turno'])

            # Aplicar bordes y alineación a celdas fijas
            for col in range(1, 5):
                cell = ws.cell(row=row_num, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            # Llenar datos de asistencia por día
            for date_obj, col_num in date_columns.items():
                cell_data = self._get_cell_data(
                    numero_empleado,
                    date_obj,
                    checadas_por_empleado,
                    info,
                    dias_con_checadas[date_obj],
                    descanso_dates,
                    incapacidad_dates,
                    public_holiday_dates
                )
                
                cell = ws.cell(row=row_num, column=col_num, value=cell_data['text'])
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                
                # Aplicar color de fondo
                if cell_data['color']:
                    cell.fill = PatternFill(start_color=cell_data['color'], end_color=cell_data['color'], fill_type='solid')
                
                # Aplicar color de fuente
                if cell_data['font_color']:
                    cell.font = Font(color=cell_data['font_color'], bold=cell_data['bold'])

            row_num += 1

        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        for col in range(5, col_num):
            ws.column_dimensions[get_column_letter(col)].width = 25

        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Crear attachment
        filename = f"Reporte_Asistencia_Importado_{self.date_from}_{self.date_to}.xlsx"
        file_data = base64.b64encode(output.getvalue())
        
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'datas': file_data,
            'type': 'binary',
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        # Retornar acción para descargar
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def _get_special_day_maps(self, employee_info, date_list, company_tz):
        """Construye mapas de fechas por empleado para Descanso/Incapacidad y festivos globales."""
        descanso_dates = {emp_num: set() for emp_num in employee_info.keys()}
        incapacidad_dates = {emp_num: set() for emp_num in employee_info.keys()}
        public_holiday_dates = set()

        if not date_list:
            return descanso_dates, incapacidad_dates, public_holiday_dates

        employee_ids = [
            info['employee_record'].id
            for info in employee_info.values()
            if info.get('employee_record')
        ]
        emp_id_to_num = {
            info['employee_record'].id: emp_num
            for emp_num, info in employee_info.items()
            if info.get('employee_record')
        }

        range_start_local = company_tz.localize(datetime.combine(self.date_from, time.min))
        range_end_local = company_tz.localize(datetime.combine(self.date_to, time.max))
        range_start_utc_str = fields.Datetime.to_string(range_start_local.astimezone(pytz.utc))
        range_end_utc_str = fields.Datetime.to_string(range_end_local.astimezone(pytz.utc))

        date_set = set(date_list)

        global_holidays = self.env['resource.calendar.leaves'].search([
            ('resource_id', '=', False),
            ('date_from', '<=', range_end_utc_str),
            ('date_to', '>=', range_start_utc_str),
        ])

        for holiday in global_holidays:
            start_local = holiday.date_from.astimezone(company_tz).date()
            end_local = holiday.date_to.astimezone(company_tz).date()
            current = start_local
            while current <= end_local:
                if current in date_set:
                    public_holiday_dates.add(current)
                current += timedelta(days=1)

        if employee_ids:
            approved_leaves = self.env['hr.leave'].sudo().search([
                ('employee_id', 'in', employee_ids),
                ('state', '=', 'validate'),
                ('date_from', '<=', range_end_utc_str),
                ('date_to', '>=', range_start_utc_str),
            ])

            for leave in approved_leaves:
                emp_num = emp_id_to_num.get(leave.employee_id.id)
                if not emp_num:
                    continue

                leave_name = (leave.holiday_status_id.name or '').strip().lower()
                is_descanso = leave_name == 'descanso'
                is_incapacidad = leave_name == 'incapacidad'

                if not is_descanso and not is_incapacidad:
                    continue

                leave_start_local = leave.date_from.astimezone(company_tz).date()
                leave_end_local = leave.date_to.astimezone(company_tz).date()

                current = leave_start_local
                while current <= leave_end_local:
                    if current in date_set:
                        if is_descanso:
                            descanso_dates.setdefault(emp_num, set()).add(current)
                        if is_incapacidad:
                            incapacidad_dates.setdefault(emp_num, set()).add(current)
                    current += timedelta(days=1)

            descanso_statuses = ['leave_day_off', 'Descanso', 'descanso']
            incapacidad_statuses = ['leave_sickness', 'leave_sickness_paid', 'Incapacidad', 'incapacidad']

            attendance_special_days = self.env['hr.attendance'].sudo().search([
                ('employee_id', 'in', employee_ids),
                ('check_in', '>=', range_start_utc_str),
                ('check_in', '<=', range_end_utc_str),
                ('punctuality_status', 'in', descanso_statuses + incapacidad_statuses),
            ])

            for attendance in attendance_special_days:
                emp_num = emp_id_to_num.get(attendance.employee_id.id)
                if not emp_num or not attendance.check_in:
                    continue

                day_local = attendance.check_in.astimezone(company_tz).date()
                if day_local not in date_set:
                    continue

                if attendance.punctuality_status in descanso_statuses:
                    descanso_dates.setdefault(emp_num, set()).add(day_local)
                if attendance.punctuality_status in incapacidad_statuses:
                    incapacidad_dates.setdefault(emp_num, set()).add(day_local)

        return descanso_dates, incapacidad_dates, public_holiday_dates

    def _get_cell_data(
        self,
        numero_empleado,
        fecha,
        checadas_por_empleado,
        employee_info,
        empleados_con_checadas_en_dia,
        descanso_dates,
        incapacidad_dates,
        public_holiday_dates,
    ):
        """
        Determina qué mostrar en la celda y su color.
        
        Lógica:
        - Verde: Más de 8 horas entre primera y última checada
        - Amarillo: Solo una checada o menos de 8 horas
        - "Descanso": Si existe en hr.leave/hr.attendance o es día no laboral del turno
        - "Pendiente de comprobar" (azul claro): Otros días sin checada cuando otros sí tienen
        """
        employee_record = employee_info.get('employee_record')
        es_domingo = fecha.weekday() == 6
        es_dia_laboral_turno = self._is_scheduled_workday(employee_record, fecha)
        es_descanso_por_turno = es_dia_laboral_turno is False
        es_descanso_registrado = fecha in descanso_dates.get(numero_empleado, set())
        es_incapacidad = fecha in incapacidad_dates.get(numero_empleado, set())
        es_festivo = fecha in public_holiday_dates

        if es_descanso_registrado and not es_incapacidad and not es_festivo:
            return {
                'text': 'Descanso',
                'color': None,
                'font_color': '808080',  # Gris
                'bold': False
            }
        
        # Verificar si el empleado tiene checadas ese día
        tiene_checadas = numero_empleado in checadas_por_empleado and fecha in checadas_por_empleado[numero_empleado]
        
        if tiene_checadas:
            checadas = checadas_por_empleado[numero_empleado][fecha]
            primera_checada = checadas[0]
            ultima_checada = checadas[-1]
            
            # Calcular diferencia en horas
            diferencia = (ultima_checada - primera_checada).total_seconds() / 3600
            
            # Formatear horas
            primera_str = primera_checada.strftime("%H:%M:%S")
            ultima_str = ultima_checada.strftime("%H:%M:%S")
            
            if len(checadas) == 1:
                # Solo una checada → Amarillo
                return {
                    'text': primera_str,
                    'color': 'FFFF00',  # Amarillo
                    'font_color': '000000',  # Negro
                    'bold': False
                }
            elif diferencia >= 8:
                # Más de 8 horas → Verde
                text = f"{primera_str} - {ultima_str}"
                return {
                    'text': text,
                    'color': '00B050',  # Verde
                    'font_color': 'FFFFFF',  # Blanco
                    'bold': False
                }
            else:
                # Menos de 8 horas → Amarillo
                text = f"{primera_str} - {ultima_str}"
                return {
                    'text': text,
                    'color': 'FFFF00',  # Amarillo
                    'font_color': '000000',  # Negro
                    'bold': False
                }
        else:
            # No tiene checadas
            if (es_descanso_por_turno or (es_dia_laboral_turno is None and es_domingo)) and not es_incapacidad and not es_festivo:
                # Día de descanso por turno. Domingo queda como fallback si no hay turno.
                return {
                    'text': 'Descanso',
                    'color': None,
                    'font_color': '808080',  # Gris
                    'bold': False
                }
            else:
                # Otro día sin checada
                # Si hay otros empleados con checadas ese día (es probable que sea día laboral)
                if empleados_con_checadas_en_dia > 0:
                    return {
                        'text': 'Pendiente de comprobar',
                        'color': 'ADD8E6',  # Azul claro
                        'font_color': '000000',  # Negro
                        'bold': False
                    }
                else:
                    # Nadie tiene checadas ese día, probablemente no laboral
                    return {
                        'text': '',
                        'color': None,
                        'font_color': None,
                        'bold': False
                    }

    def _is_scheduled_workday(self, employee, target_date):
        """Retorna True/False según el turno del empleado, o None si no se puede determinar."""
        if not employee or not employee.turno_id:
            return None

        weekday_to_field = {
            0: 'work_monday',
            1: 'work_tuesday',
            2: 'work_wednesday',
            3: 'work_thursday',
            4: 'work_friday',
            5: 'work_saturday',
            6: 'work_sunday',
        }

        work_field = weekday_to_field.get(target_date.weekday())
        if not work_field:
            return None

        return bool(getattr(employee.turno_id, work_field, False))
