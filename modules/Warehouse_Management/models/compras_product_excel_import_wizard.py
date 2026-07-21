import base64
from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from odoo import _, fields, models
from odoo.exceptions import ValidationError


class ComprasProductExcelImportWizard(models.TransientModel):
    _name = 'compras.product.excel.import.wizard'
    _description = 'Importar Consolidado desde Excel'

    file_data = fields.Binary(string='Archivo Excel')
    file_name = fields.Char(string='Nombre de archivo')
    template_file_data = fields.Binary(string='Plantilla Excel', readonly=True, attachment=False)
    template_file_name = fields.Char(string='Nombre de plantilla', readonly=True)

    _EXPECTED_HEADERS = [
        ('code', {'codigo'}),
        ('name', {'descripcion'}),
        ('location', {'locacion'}),
        ('model_name', {'modelo'}),
        ('serial', {'serial'}),
        ('manufacturer', {'fabricante'}),
        ('classification', {'clasificacion'}),
        ('category_id', {'categoria'}),
        ('subcategory_id', {'sub categoria', 'subcategoria'}),
        ('subcategory_2', {'sub categoria 2', 'subcategoria 2'}),
        ('repair_line_id', {'equipo linea a', 'equipo - linea a reparar', 'equipo - linea a'}),
        ('expiration_date', {'fecha de cad', 'fecha de caducidad'}),
        ('qty_process', {'qty proceso', 'qty - proceso', 'qty-proceso'}),
        ('total_equipment', {'total equipos', 'total'}),
        ('total_qty_process', {'total qty proceso', 'total qty procesos', 'total qty'}),
        ('frequency_use_days', {'frecuencia uso', 'frecuencia uso dias'}),
        ('unit_id', {'udm', 'unidad'}),
        ('unit_price', {'costo unitario', 'precio unitario'}),
        ('total_cost', {'costo total'}),
        ('currency_id', {'moneda'}),
        ('qty_on_hand', {'inventario'}),
        ('coverage_days', {'cobertura'}),
        ('min_qty', {'min'}),
        ('max_qty', {'max'}),
        ('reorder_point', {'punto reorden', 'punto reorder'}),
        ('lead_time_days', {'tiempo entrega', 'tiempo de entrega', 'tiempo de entrega dias'}),
        ('purchase_type', {'tipo compra', 'tipo de compra'}),
        ('vendor_id', {'proveedor principal', 'proveeder principal'}),
        ('monthly_budget', {'presupuesto mensual'}),
    ]

    def action_import_excel(self):
        self.ensure_one()
        if not self.file_data:
            raise ValidationError(_('Debes seleccionar un archivo Excel.'))

        rows = self._read_excel_rows()
        if not rows:
            raise ValidationError(_('El archivo no contiene datos para importar.'))

        header_row = rows[0]
        self._validate_header_order(header_row)

        product_model = self.env['compras.product']
        company = self.env.company
        created_count = 0
        updated_count = 0
        skipped_rows_location = []
        rows_without_location = []
        rows_without_inventory_adjustment = []
        row_errors = []

        for line_number, row in enumerate(rows[1:], start=2):
            if self._is_empty_row(row):
                continue
            try:
                code = self._as_text(self._cell_value(row, 0))
                if not code:
                    raise ValidationError(_('El Código es obligatorio.'))

                location_name = self._as_text(self._cell_value(row, 2))
                warehouse = False
                location_record = False
                if location_name:
                    warehouse, location_record = self._find_or_create_location_by_code(location_name, company)
                    if not (warehouse and location_record):
                        skipped_rows_location.append(line_number)
                        continue
                else:
                    rows_without_location.append(line_number)

                qty_on_hand = self._to_float(self._cell_value(row, 20), _('Inventario'))

                vals = self._build_product_vals_from_row(
                    row,
                    company,
                    warehouse=warehouse,
                    location_record=location_record,
                )
                product = product_model.search([
                    ('company_id', '=', company.id),
                    ('code', '=', code),
                ], limit=1)

                if product:
                    product.write(vals)
                    if qty_on_hand is not None:
                        if warehouse and location_record:
                            product.write({'qty_on_hand': qty_on_hand})
                        else:
                            rows_without_inventory_adjustment.append(line_number)
                    updated_count += 1
                else:
                    vals.setdefault('code', code)
                    vals.setdefault('company_id', company.id)
                    if not vals.get('name'):
                        raise ValidationError(_('La Descripción es obligatoria para crear un nuevo registro.'))
                    created_product = product_model.create(vals)
                    if qty_on_hand is not None:
                        if warehouse and location_record:
                            created_product.write({'qty_on_hand': qty_on_hand})
                        else:
                            rows_without_inventory_adjustment.append(line_number)
                    created_count += 1
            except ValidationError as validation_error:
                row_errors.append(_('Fila %(line)s: %(message)s') % {
                    'line': line_number,
                    'message': validation_error.args[0],
                })

        if row_errors:
            raise ValidationError('\n'.join(row_errors))

        summary = _('%(created)s creados, %(updated)s actualizados.') % {
            'created': created_count,
            'updated': updated_count,
        }
        if skipped_rows_location:
            summary += ' ' + _(
                'Se omitieron %(count)s fila(s) porque la locación no coincide con un almacén válido de la empresa: %(rows)s.'
            ) % {
                'count': len(skipped_rows_location),
                'rows': ', '.join(str(line) for line in skipped_rows_location),
            }
        if rows_without_location:
            summary += ' ' + _(
                'Se detectaron %(count)s fila(s) sin locación; se importaron sin asignar almacén/locación y sin ajustar inventario: %(rows)s.'
            ) % {
                'count': len(rows_without_location),
                'rows': ', '.join(str(line) for line in rows_without_location),
            }
        if rows_without_inventory_adjustment:
            summary += ' ' + _(
                'No se ajustó inventario en %(count)s fila(s) por falta de almacén/locación válida: %(rows)s.'
            ) % {
                'count': len(rows_without_inventory_adjustment),
                'rows': ', '.join(str(line) for line in rows_without_inventory_adjustment),
            }
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Importación finalizada'),
                'message': summary,
                'type': 'success',
                'sticky': False,
                'next': {'type': 'ir.actions.act_window_close'},
            },
        }

    def action_download_template(self):
        self.ensure_one()

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Plantilla Consolidado'

        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        example_fill = PatternFill(start_color='EAF2F8', end_color='EAF2F8', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        example_font = Font(italic=True, color='666666')
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        headers = [
            'Código',
            'Descripción',
            'Locación',
            'Modelo',
            'Serial',
            'Fabricante',
            'Clasificación',
            'Categoría',
            'Sub categoría',
            'Sub categoría 2',
            'Equipo / línea a reparar',
            'Fecha de caducidad',
            'Qty proceso',
            'Total equipos',
            'Total qty proceso',
            'Frecuencia uso días',
            'UDM',
            'Costo unitario',
            'Costo total',
            'Moneda',
            'Inventario',
            'Cobertura',
            'Min',
            'Max',
            'Punto reorden',
            'Tiempo de entrega dias',
            'Tipo compra',
            'Proveedor principal',
            'Presupuesto mensual',
        ]
        sample_row = [
            'COD-001',
            'Bomba centrífuga industrial',
            'Planta 1',
            'Modelo X100',
            'SN-0001',
            'Acme',
            'Equipo',
            'Bombas',
            'Bombas principales',
            'Serie A',
            'Mantenimiento',
            '2026-12-31',
            12,
            3,
            15,
            30,
            'Unidad',
            1250.50,
            3751.50,
            'MXN',
            8,
            20,
            5,
            25,
            10,
            7,
            'local',
            'Proveedor Demo SA de CV',
            50000,
        ]

        for column_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=column_index, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border

        for column_index, value in enumerate(sample_row, start=1):
            cell = sheet.cell(row=2, column=column_index, value=value)
            cell.fill = example_fill
            cell.font = example_font
            cell.alignment = left_alignment
            cell.border = border

        sheet.freeze_panes = 'A2'
        sheet.auto_filter.ref = sheet.dimensions

        column_widths = {
            'A': 16,
            'B': 28,
            'C': 18,
            'D': 20,
            'E': 16,
            'F': 18,
            'G': 18,
            'H': 20,
            'I': 20,
            'J': 18,
            'K': 24,
            'L': 18,
            'M': 14,
            'N': 14,
            'O': 16,
            'P': 16,
            'Q': 14,
            'R': 16,
            'S': 16,
            'T': 12,
            'U': 14,
            'V': 12,
            'W': 12,
            'X': 12,
            'Y': 14,
            'Z': 14,
            'AA': 14,
            'AB': 22,
            'AC': 16,
        }
        for column_name, width in column_widths.items():
            sheet.column_dimensions[column_name].width = width

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = 'plantilla_consolidado.xlsx'
        self.write({
            'template_file_data': base64.b64encode(output.getvalue()),
            'template_file_name': filename,
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/?model=compras.product.excel.import.wizard&id={self.id}&field=template_file_data&filename_field=template_file_name&download=true',
            'target': 'self',
        }

    def _read_excel_rows(self):
        try:
            workbook = load_workbook(filename=BytesIO(base64.b64decode(self.file_data)), data_only=True)
        except Exception as exc:
            raise ValidationError(_('No se pudo leer el archivo Excel: %s') % str(exc)) from exc

        sheet = workbook.active
        return [list(row) for row in sheet.iter_rows(values_only=True)]

    def _validate_header_order(self, header_row):
        if len(header_row) < len(self._EXPECTED_HEADERS):
            raise ValidationError(_('El archivo debe contener al menos %(count)s columnas en el orden definido.') % {
                'count': len(self._EXPECTED_HEADERS),
            })

        normalized_headers = [self._normalize_header_name(value) for value in header_row]

        for index, (field_name, allowed_names) in enumerate(self._EXPECTED_HEADERS):
            normalized_allowed = {self._normalize_header_name(name) for name in allowed_names}
            if normalized_headers[index] not in normalized_allowed:
                expected_values = ', '.join(sorted(allowed_names))
                raise ValidationError(_(
                    'La columna %(position)s no coincide. Se esperaba: %(expected)s. Valor recibido: %(received)s'
                ) % {
                    'position': index + 1,
                    'expected': expected_values,
                    'received': header_row[index] or '',
                })

    def _build_product_vals_from_row(
        self,
        row,
        company,
        warehouse=False,
        location_record=False,
    ):
        vals = {}

        name = self._as_text(self._cell_value(row, 1))
        if name:
            vals['name'] = name

        if warehouse and location_record:
            vals['inventory_warehouse_id'] = warehouse.id
            vals['inventory_location_id'] = location_record.id
            vals['location'] = location_record.name

        model_name = self._as_text(self._cell_value(row, 3))
        if model_name:
            vals['model_name'] = model_name

        serial = self._as_text(self._cell_value(row, 4))
        if serial:
            vals['serial'] = serial

        manufacturer = self._as_text(self._cell_value(row, 5))
        if manufacturer:
            vals['manufacturer'] = manufacturer

        classification = self._as_text(self._cell_value(row, 6))
        if classification:
            vals['classification'] = classification

        category_name = self._as_text(self._cell_value(row, 7))
        category = self._find_or_create_category(category_name)
        if category:
            vals['category_id'] = category.id

        subcategory_name = self._as_text(self._cell_value(row, 8))
        subcategory = self._find_or_create_subcategory(subcategory_name, category)
        if subcategory:
            vals['subcategory_id'] = subcategory.id

        subcategory_2 = self._as_text(self._cell_value(row, 9))
        if subcategory_2:
            vals['subcategory_2'] = subcategory_2

        repair_line_name = self._as_text(self._cell_value(row, 10))
        repair_line = self._find_or_create_repair_line(repair_line_name)
        if repair_line:
            vals['repair_line_id'] = repair_line.id

        expiration_date = self._to_date(self._cell_value(row, 11))
        if expiration_date:
            vals['expiration_date'] = expiration_date

        qty_process = self._to_float(self._cell_value(row, 12), _('Qty - Proceso'))
        if qty_process is not None:
            vals['qty_process'] = qty_process

        total_equipment = self._to_float(self._cell_value(row, 13), _('Total Equipos'))
        if total_equipment is not None:
            vals['total_equipment'] = total_equipment

        frequency_use_days = self._to_float(self._cell_value(row, 15), _('Frecuencia Uso'))
        if frequency_use_days is not None:
            vals['frequency_use_days'] = frequency_use_days

        unit_name = self._as_text(self._cell_value(row, 16))
        unit = self._find_uom(unit_name)
        if unit:
            vals['unit_id'] = unit.id

        unit_price = self._to_float(self._cell_value(row, 17), _('Costo Unitario'))
        if unit_price is not None:
            vals['unit_price'] = unit_price

        currency_name = self._as_text(self._cell_value(row, 19))
        currency = self._find_currency(currency_name)
        if currency:
            vals['currency_id'] = currency.id

        min_qty = self._to_float(self._cell_value(row, 22), _('MIN'))
        if min_qty is not None:
            vals['min_qty'] = min_qty

        lead_time_days = self._to_float(self._cell_value(row, 25), _('Tiempo Entrega'))
        if lead_time_days is not None:
            vals['lead_time_days'] = lead_time_days

        purchase_type = self._to_purchase_type(self._cell_value(row, 26))
        if purchase_type:
            vals['purchase_type'] = purchase_type

        vendor_name = self._as_text(self._cell_value(row, 27))
        vendor = self._find_or_create_vendor(vendor_name, company)
        if vendor:
            vals['vendor_id'] = vendor.id

        return vals

    def _find_or_create_category(self, name):
        if not name:
            return False
        category_model = self.env['compras.product.category']
        category = category_model.search([('name', '=', name)], limit=1)
        return category or category_model.create({'name': name})

    def _find_or_create_subcategory(self, name, category):
        if not name or not category:
            return False
        subcategory_model = self.env['compras.product.subcategory']
        subcategory = subcategory_model.search([
            ('name', '=', name),
            ('category_id', '=', category.id),
        ], limit=1)
        return subcategory or subcategory_model.create({'name': name, 'category_id': category.id})

    def _find_or_create_repair_line(self, name):
        if not name:
            return False
        repair_line_model = self.env['compras.repair.line']
        repair_line = repair_line_model.search([('name', '=', name)], limit=1)
        return repair_line or repair_line_model.create({'name': name})

    def _find_uom(self, unit_name):
        if not unit_name:
            return False
        uom_model = self.env['uom.uom']
        unit = uom_model.search([('name', '=', unit_name)], limit=1)
        if not unit:
            unit = uom_model.search([('name', 'ilike', unit_name)], limit=1)
        return unit

    def _find_currency(self, currency_name):
        if not currency_name:
            return False
        currency_model = self.env['res.currency']
        currency = currency_model.search([('name', '=', currency_name)], limit=1)
        if not currency:
            currency = currency_model.search([('name', 'ilike', currency_name)], limit=1)
        if not currency:
            currency = currency_model.search([('symbol', '=', currency_name)], limit=1)
        if not currency:
            raise ValidationError(_('No se encontró la moneda: %s') % currency_name)
        return currency

    def _find_or_create_vendor(self, vendor_name, company):
        if not vendor_name:
            return False

        vendor_in_products = self.env['compras.product'].search([
            ('company_id', '=', company.id),
            ('vendor_id', '!=', False),
            ('vendor_id.name', '=ilike', vendor_name),
        ], limit=1)
        if vendor_in_products and vendor_in_products.vendor_id:
            return vendor_in_products.vendor_id

        partner_model = self.env['res.partner']
        vendor = partner_model.search([('name', '=', vendor_name)], limit=1)
        if not vendor:
            vendor = partner_model.search([('name', 'ilike', vendor_name)], limit=1)
        if not vendor:
            vendor = partner_model.create({
                'name': vendor_name,
                'supplier_rank': 1,
                'company_type': 'company',
            })
        elif vendor.supplier_rank <= 0:
            vendor.supplier_rank = 1
        return vendor

    def _find_or_create_location_by_code(self, location_name, company):
        if not location_name:
            return (False, False)

        normalized_location = self._as_text(location_name)
        if not normalized_location:
            return (False, False)

        warehouse_code = normalized_location.split('-', 1)[0].strip()
        if not warehouse_code:
            return (False, False)

        warehouse_model = self.env['compras.warehouse']
        warehouse = warehouse_model.search([
            ('company_id', '=', company.id),
            ('code', '=', warehouse_code),
        ], limit=1)
        if not warehouse:
            warehouse = warehouse_model.search([
                ('company_id', '=', company.id),
                ('code', '=ilike', warehouse_code),
            ], limit=1)
        if not warehouse:
            return (False, False)

        location_model = self.env['compras.warehouse.location']
        location = location_model.search([
            ('warehouse_id', '=', warehouse.id),
            ('name', '=', normalized_location),
        ], limit=1)
        if not location:
            location = location_model.search([
                ('warehouse_id', '=', warehouse.id),
                ('name', '=ilike', normalized_location),
            ], limit=1)
        if not location:
            location = location_model.create({
                'name': normalized_location,
                'warehouse_id': warehouse.id,
            })

        return (warehouse, location)

    def _to_purchase_type(self, value):
        purchase_type = self._normalize_header_name(value)
        if not purchase_type:
            return False
        if purchase_type in ('local',):
            return 'local'
        if purchase_type in ('internacional', 'international'):
            return 'internacional'
        raise ValidationError(_('Tipo de compra inválido: %s') % value)

    def _to_float(self, value, field_label):
        if value in (None, ''):
            return None
        if isinstance(value, (int, float)):
            return float(value)

        text_value = str(value).strip().replace(',', '')
        if not text_value:
            return None

        try:
            return float(text_value)
        except ValueError as exc:
            raise ValidationError(_('%(field)s no es un número válido: %(value)s') % {
                'field': field_label,
                'value': value,
            }) from exc

    def _to_date(self, value):
        if value in (None, ''):
            return False
        if isinstance(value, datetime):
            return fields.Date.to_string(value.date())
        if isinstance(value, date):
            return fields.Date.to_string(value)

        text_value = str(value).strip()
        if not text_value:
            return False

        parsed_date = fields.Date.to_date(text_value)
        if not parsed_date:
            raise ValidationError(_('Formato de fecha inválido: %s') % value)
        return fields.Date.to_string(parsed_date)

    def _normalize_header_name(self, value):
        normalized = self._as_text(value).lower()
        replacements = {
            'á': 'a',
            'é': 'e',
            'í': 'i',
            'ó': 'o',
            'ú': 'u',
            '.': ' ',
            '_': ' ',
            '/': ' ',
            '-': ' ',
        }
        for old_value, new_value in replacements.items():
            normalized = normalized.replace(old_value, new_value)
        return ' '.join(normalized.split())

    def _cell_value(self, row, index):
        if index >= len(row):
            return None
        return row[index]

    def _is_empty_row(self, row):
        return not any(self._as_text(value) for value in row)

    def _as_text(self, value):
        if value is None:
            return ''
        return str(value).strip()
